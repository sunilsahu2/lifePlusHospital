from flask import Flask, request, jsonify, render_template, send_from_directory, send_file, session
from flask_cors import CORS
from pymongo import MongoClient
from bson import ObjectId
from datetime import datetime, timedelta
import urllib.parse
import logging
import os
import uuid
import hashlib
from werkzeug.utils import secure_filename
from functools import wraps
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import io
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch

# Configure logging
logging.basicConfig(
    level=logging.DEBUG,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('app.log'),
        logging.StreamHandler()
    ]
)

app = Flask(__name__)
app.secret_key = os.getenv('SECRET_KEY', 'your-secret-key-change-in-production')
app.config['UPLOAD_FOLDER'] = os.path.join(os.getcwd(), 'static', 'uploads')
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'
app.config['SESSION_COOKIE_SECURE'] = False # Set to True in production with HTTPS
CORS(app, supports_credentials=True)

# MongoDB Connection Configuration
MONGODB_URI = os.getenv('MONGODB_URI')

if not MONGODB_URI:
    MONGODB_USERNAME = 'sunilsahu'
    MONGODB_PASSWORD = 'sokwer-pubgux-poxxE3'
    MONGODB_CLUSTER = 'cluster0.qufitms.mongodb.net'
    MONGODB_DB_NAME = 'hospital_management'
    encoded_password = urllib.parse.quote_plus(MONGODB_PASSWORD)
    MONGODB_URI = f'mongodb+srv://{MONGODB_USERNAME}:{encoded_password}@{MONGODB_CLUSTER}/{MONGODB_DB_NAME}?retryWrites=true&w=majority'
else:
    # Extract DB name from URI if possible, or use default
    MONGODB_DB_NAME = os.getenv('MONGODB_DB_NAME', 'hospital_management')

# Connect to MongoDB
try:
    client = MongoClient(MONGODB_URI, serverSelectionTimeoutMS=10000)
    client.server_info()  # Test connection
    db = client[MONGODB_DB_NAME]
    logging.info(f"✓ Successfully connected to MongoDB Atlas database '{MONGODB_DB_NAME}'")
except Exception as e:
    logging.error(f"✗ MongoDB connection error: {e}")
    raise

# Helper function to convert ObjectId to string
def serialize_doc(doc):
    if doc is None:
        return None
    if isinstance(doc, list):
        return [serialize_doc(item) for item in doc]
    if isinstance(doc, dict):
        doc = doc.copy()
        if '_id' in doc:
            doc['id'] = str(doc.pop('_id'))
        # Convert any ObjectId fields to strings
        for key, value in doc.items():
            if isinstance(value, ObjectId):
                doc[key] = str(value)
        return doc
    return doc

# Helper function to parse ObjectId
def parse_object_id(id_str):
    try:
        if not id_str:
            return None
        return ObjectId(id_str)
    except:
        return None

def is_case_closed(case_id):
    """Check if a case is closed. Returns True if closed, False otherwise."""
    if not case_id:
        return False
    # Ensure it's an ObjectId for the query
    from bson.objectid import ObjectId as BsonObjectId
    if not isinstance(case_id, BsonObjectId):
        case_id = parse_object_id(str(case_id))
    if not case_id:
        return False
    case = db.cases.find_one({'_id': case_id}, {'status': 1})
    return case and case.get('status') == 'closed'

def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            # Fallback to header for robustness
            user_id_header = request.headers.get('X-User-Id')
            if user_id_header:
                try:
                    user = db.users.find_one({'_id': parse_object_id(user_id_header)})
                    if user:
                        # Populate session for this request context
                        session['user_id'] = str(user['_id'])
                        session['role'] = user.get('role', 'staff')
                        session['username'] = user.get('username')
                        logging.debug(f"Header Auth Success. Set session role: {session.get('role')}")
                    else:
                        return jsonify({'error': 'Invalid User ID header'}), 401
                except:
                     return jsonify({'error': 'Invalid User ID format'}), 401
            else:
                return jsonify({'error': 'Authentication required'}), 401
        return f(*args, **kwargs)
    return decorated_function

# ==================== ROUTES ====================

@app.route('/')
def index():
    return render_template('index.html')

@app.before_request
def log_request_info():
    logging.debug('Headers: %s', request.headers)
    logging.debug('Cookies: %s', request.cookies)
    logging.debug('Session: %s', session)
    if 'user_id' in session:
        logging.debug(f"User ID in session: {session['user_id']}")
    else:
        logging.debug("No User ID in session")


@app.route('/static/uploads/prescriptions/<filename>')
# Trigger reload
def uploaded_file(filename):
    return send_from_directory(os.path.join(app.root_path, 'static', 'uploads', 'prescriptions'), filename)

@app.route('/static/manifest.json')
def manifest():
    return send_from_directory(os.path.join(app.root_path, 'static'), 'manifest.json', mimetype='application/json')

@app.route('/static/service-worker.js')
def service_worker():
    return send_from_directory(os.path.join(app.root_path, 'static'), 'service-worker.js', mimetype='application/javascript')

# ==================== DOCTORS API ====================

@app.route('/api/doctors', methods=['GET'])
def get_doctors():
    try:
        page = int(request.args.get('page', 1))
        limit = int(request.args.get('limit', 10))
        search = request.args.get('search', '').strip()
        skip = (page - 1) * limit
        
        # Build query for search - by default only show active doctors
        query = {'$or': [{'isActive': True}, {'isActive': {'$exists': False}}]}  # Show active doctors (isActive is True or doesn't exist)
        if search:
            search_conditions = [
                {'name': {'$regex': search, '$options': 'i'}},
                {'specialization': {'$regex': search, '$options': 'i'}},
                {'phone': {'$regex': search, '$options': 'i'}},
                {'email': {'$regex': search, '$options': 'i'}}
            ]
            query = {
                '$and': [
                    {'$or': [{'isActive': True}, {'isActive': {'$exists': False}}]},
                    {'$or': search_conditions}
                ]
            }
        
        # Get total count for pagination
        total = db.doctors.count_documents(query)
        
        # Get paginated results
        doctors = list(db.doctors.find(query).sort('created_at', -1).skip(skip).limit(limit))
        
        return jsonify({
            'doctors': serialize_doc(doctors),
            'total': total,
            'page': page,
            'limit': limit
        })
    except Exception as e:
        logging.error(f"Error getting doctors: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/doctors/<id>', methods=['GET'])
def get_doctor(id):
    try:
        doctor = db.doctors.find_one({'_id': parse_object_id(id)})
        if doctor:
            return jsonify(serialize_doc(doctor))
        return jsonify({'error': 'Doctor not found'}), 404
    except Exception as e:
        logging.error(f"Error getting doctor: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/doctors', methods=['POST'])
def create_doctor():
    try:
        data = request.get_json()
        
        # Validate required fields
        if not data.get('name'):
            return jsonify({'error': 'Name is required'}), 400
        
        # Set defaults
        data['created_at'] = datetime.now()
        if 'isActive' not in data:
            data['isActive'] = True  # Set as active by default
        
        # Handle isInhouse flag
        if 'isInhouse' not in data:
            data['isInhouse'] = False
        else:
            # Ensure it's boolean if sent as string from some source
            if isinstance(data['isInhouse'], str):
                data['isInhouse'] = data['isInhouse'].lower() == 'true'
        
        result = db.doctors.insert_one(data)
        return jsonify({'id': str(result.inserted_id), 'message': 'Doctor created successfully'}), 201
    except Exception as e:
        logging.error(f"Error creating doctor: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/doctors/<id>', methods=['PUT'])
def update_doctor(id):
    try:
        data = request.get_json()
        doctor_id = parse_object_id(id)
        
        # Check if doctor exists
        existing_doctor = db.doctors.find_one({'_id': doctor_id})
        if not existing_doctor:
            return jsonify({'error': 'Doctor not found'}), 404
        
        # Don't allow updating isActive through PUT (use DELETE endpoint for deactivation)
        if 'isActive' in data:
            del data['isActive']
        
        # Handle isInhouse flag if present
        if 'isInhouse' in data:
            if isinstance(data['isInhouse'], str):
                data['isInhouse'] = data['isInhouse'].lower() == 'true'
        
        data['updated_at'] = datetime.now()
        result = db.doctors.update_one({'_id': doctor_id}, {'$set': data})
        if result.modified_count or result.matched_count:
            return jsonify({'message': 'Doctor updated successfully'})
        return jsonify({'message': 'Doctor updated successfully'})  # Even if no fields changed
    except Exception as e:
        logging.error(f"Error updating doctor: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/doctors/<id>', methods=['DELETE'])
def delete_doctor(id):
    try:
        result = db.doctors.update_one(
            {'_id': parse_object_id(id)},
            {'$set': {'isActive': False, 'deactivated_at': datetime.now()}}
        )
        if result.modified_count:
            return jsonify({'message': 'Doctor deactivated successfully'})
        return jsonify({'error': 'Doctor not found'}), 404
    except Exception as e:
        logging.error(f"Error deactivating doctor: {e}")
        return jsonify({'error': str(e)}), 500

# ==================== PATIENTS API ====================

@app.route('/api/patients', methods=['GET'])
def get_patients():
    try:
        page = int(request.args.get('page', 1))
        limit = int(request.args.get('limit', 10))
        search = request.args.get('search', '').strip()
        skip = (page - 1) * limit
        
        # Build query for search
        query = {}
        if search:
            search_conditions = [
                {'name': {'$regex': search, '$options': 'i'}},
                {'phone': {'$regex': search, '$options': 'i'}},
                {'email': {'$regex': search, '$options': 'i'}},
                {'address': {'$regex': search, '$options': 'i'}}
            ]
            query = {'$or': search_conditions}
        
        # Get total count for pagination
        total = db.patients.count_documents(query)
        
        # Get paginated results
        patients = list(db.patients.find(query).sort('created_at', -1).skip(skip).limit(limit))
        
        return jsonify({
            'patients': serialize_doc(patients),
            'total': total,
            'page': page,
            'limit': limit
        })
    except Exception as e:
        logging.error(f"Error getting patients: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/patients/<id>', methods=['GET'])
def get_patient(id):
    try:
        patient = db.patients.find_one({'_id': parse_object_id(id)})
        if patient:
            return jsonify(serialize_doc(patient))
        return jsonify({'error': 'Patient not found'}), 404
    except Exception as e:
        logging.error(f"Error getting patient: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/patients', methods=['POST'])
def create_patient():
    try:
        data = request.get_json()
        data['created_at'] = datetime.now()
        result = db.patients.insert_one(data)
        return jsonify({'id': str(result.inserted_id), 'message': 'Patient created successfully'}), 201
    except Exception as e:
        logging.error(f"Error creating patient: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/patients/<id>', methods=['PUT'])
def update_patient(id):
    try:
        data = request.get_json()
        data['updated_at'] = datetime.now()
        result = db.patients.update_one({'_id': parse_object_id(id)}, {'$set': data})
        if result.modified_count:
            return jsonify({'message': 'Patient updated successfully'})
        return jsonify({'error': 'Patient not found'}), 404
    except Exception as e:
        logging.error(f"Error updating patient: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/patients/<id>', methods=['DELETE'])
def delete_patient(id):
    try:
        result = db.patients.delete_one({'_id': parse_object_id(id)})
        if result.deleted_count:
            return jsonify({'message': 'Patient deleted successfully'})
        return jsonify({'error': 'Patient not found'}), 404
    except Exception as e:
        logging.error(f"Error deleting patient: {e}")
        return jsonify({'error': str(e)}), 500

# ==================== CASES API ====================

@app.route('/api/cases', methods=['GET'])
def get_cases():
    try:
        page = int(request.args.get('page', 1))
        limit = int(request.args.get('limit', 10))
        search = request.args.get('search', '').strip()
        patient_id = request.args.get('patient_id')
        skip = (page - 1) * limit
        
        # Build query
        filters = []
        
        # Filter by patient_id
        if patient_id:
            filters.append({'patient_id': parse_object_id(patient_id)})
            
        # Filter by status
        status = request.args.get('status')
        if status:
            if status.lower() == 'open':
                filters.append({'$or': [
                    {'status': 'open'},
                    {'status': {'$exists': False}},
                    {'status': None}
                ]})
            else:
                filters.append({'status': status.lower()})
        
        # Filter by search
        if search:
            # Search by case_number (exact match preferred)
            cases_by_number = list(db.cases.find({'case_number': {'$regex': search, '$options': 'i'}}))
            case_ids = [case['_id'] for case in cases_by_number]
            
            # Also search in patients by name, phone, email
            patients = list(db.patients.find({
                '$or': [
                    {'name': {'$regex': search, '$options': 'i'}},
                    {'phone': {'$regex': search, '$options': 'i'}},
                    {'email': {'$regex': search, '$options': 'i'}}
                ]
            }))
            patient_ids = [patient['_id'] for patient in patients]
            
            search_conditions = []
            if case_ids:
                search_conditions.append({'_id': {'$in': case_ids}})
            if patient_ids:
                search_conditions.append({'patient_id': {'$in': patient_ids}})
            
            if search_conditions:
                filters.append({'$or': search_conditions})
            else:
                filters.append({'_id': {'$exists': False}})

        # Construct final query
        if len(filters) > 1:
            query = {'$and': filters}
        elif len(filters) == 1:
            query = filters[0]
        else:
            query = {}
        
        # Get total count for pagination
        total = db.cases.count_documents(query)
        
        # Aggregation Pipeline
        pipeline = [
            {'$match': query},
            {'$sort': {'created_at': -1}},
            {'$skip': skip},
            {'$limit': limit},
            
            # Lookup Patient
            {'$lookup': {
                'from': 'patients',
                'localField': 'patient_id',
                'foreignField': '_id',
                'as': 'patient_doc'
            }},
            {'$addFields': {
                'patient_name': {'$ifNull': [{'$arrayElemAt': ['$patient_doc.name', 0]}, '']}
            }},
            
            # Lookup Hospital Charges
            {'$lookup': {
                'from': 'case_charges',
                'localField': '_id',
                'foreignField': 'case_id',
                'as': 'hospital_charges_list'
            }},
            
            # Lookup Doctor Charges (Legacy)
            {'$lookup': {
                'from': 'case_doctor_charges',
                'localField': '_id',
                'foreignField': 'case_id',
                'as': 'doctor_charges_list'
            }},
            
            # Lookup Payments
            {'$lookup': {
                'from': 'payments',
                'localField': '_id',
                'foreignField': 'case_id',
                'as': 'payments_list'
            }},
            
            # Lookup Appointments (with doctor info)
            {'$lookup': {
                'from': 'appointments',
                'let': {'caseId': '$_id'},
                'pipeline': [
                    {'$match': {'$expr': {'$eq': ['$case_id', '$$caseId']}}},
                    {'$sort': {'appointment_date': 1}},
                    # Lookup doctor for appointment
                    {'$lookup': {
                        'from': 'doctors',
                        'localField': 'doctor_id',
                        'foreignField': '_id',
                        'as': 'doctor_doc'
                    }},
                    {'$addFields': {
                        'doctor_name': {'$arrayElemAt': ['$doctor_doc.name', 0]}
                    }}
                ],
                'as': 'appointments_data'
            }},
            
            # Calculate Totals
            {'$addFields': {
                'hospital_total': {'$sum': '$hospital_charges_list.total_amount'},
                'hospital_count': {'$size': '$hospital_charges_list'},
                'doctor_total': {'$sum': '$doctor_charges_list.amount'},
                'doctor_count': {'$size': '$doctor_charges_list'},
                'paid_amount': {'$sum': '$payments_list.amount'},
                'discount_val': {'$ifNull': ['$discount', 0]} 
            }},
            
            {'$addFields': {
                'charges_total': {'$add': ['$hospital_total', '$doctor_total']},
                'charges_count': {'$add': ['$hospital_count', '$doctor_count']},
                'due_amount': {'$max': [0, {'$subtract': [
                    {'$max': [0, {'$subtract': [{'$add': ['$hospital_total', '$doctor_total']}, '$discount_val']}]}, 
                    '$paid_amount'
                ]}]},
                'appointments_count': {'$size': '$appointments_data'}
            }},
            
            # Cleanup temporary fields
            {'$project': {
                'patient_doc': 0,
                'hospital_charges_list': 0,
                'doctor_charges_list': 0,
                'payments_list': 0,
                'hospital_total': 0,
                'doctor_total': 0,
                'doctor_count': 0,
                'hospital_count': 0
            }}
        ]
        
        cases = list(db.cases.aggregate(pipeline))
        
        # Post-process appointments (logic difficult to do purely in aggregation)
        now_date = datetime.now().date()
        
        for case in cases:
            # Ensure status
            if 'status' not in case:
                case['status'] = 'open'
                
            # Handle Next Appointment Logic
            appointments_data = case.pop('appointments_data', [])
            
            # Filter for display (first 5)
            # Original logic: "limit to 5 most recent for display" -> assumes sorted? 
            # Original code sorted by appointment_date 1. So "recent" might mean future ones?
            # Or just the list. Code said: case_appointments[:5]
            
            display_apts = []
            for apt in appointments_data[:5]:
                apt_display = {
                    'id': str(apt['_id']),
                    'appointment_date': apt.get('appointment_date'),
                    'appointment_time': apt.get('appointment_time'),
                    'status': apt.get('status'),
                    'doctor_name': apt.get('doctor_name', '')
                }
                display_apts.append(apt_display)
            case['appointments'] = serialize_doc(display_apts)
            
            # Find next upcoming appointment
            upcoming_appointments = [
                apt for apt in appointments_data 
                if apt.get('status') != 'Cancelled' and apt.get('appointment_date')
            ]
            
            if upcoming_appointments:
                future_appointments = []
                for apt in upcoming_appointments:
                    apt_date = apt.get('appointment_date')
                    if isinstance(apt_date, datetime):
                        apt_date_obj = apt_date.date()
                    elif isinstance(apt_date, str):
                        try:
                            apt_date_obj = datetime.fromisoformat(apt_date.split('T')[0]).date()
                        except:
                            continue
                    else:
                        continue
                        
                    if apt_date_obj >= now_date:
                        future_appointments.append(apt)
                
                if future_appointments:
                    next_apt = future_appointments[0] # Already sorted by date in pipeline
                    case['next_appointment_date'] = next_apt.get('appointment_date')
                    case['next_appointment_time'] = next_apt.get('appointment_time')
                    case['next_appointment_doctor'] = next_apt.get('doctor_name', '')

            # Convert patient_id to string loop handled by serialize_doc later, 
            # but we need to ensure patient_id field format matches legacy if manual handling needed.
            # serialize_doc will handle ObjectIds.
            
        return jsonify({
            'cases': serialize_doc(cases),
            'total': total,
            'page': page,
            'limit': limit
        })
    except Exception as e:
        logging.error(f"Error getting cases: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/cases/<id>', methods=['GET'])
def get_case(id):
    try:
        case = db.cases.find_one({'_id': parse_object_id(id)})
        if not case:
            return jsonify({'error': 'Case not found'}), 404
        
        # Populate patient
        if 'patient_id' in case:
            patient_id_obj = case['patient_id']
            # Convert patient_id to ObjectId if it's a string
            if isinstance(patient_id_obj, str):
                patient_id_obj = parse_object_id(patient_id_obj)
            if patient_id_obj:
                patient = db.patients.find_one({'_id': patient_id_obj})
                if patient:
                    case['patient'] = serialize_doc(patient)
        
        # Populate referred_by name
        if 'referred_by_id' in case and 'referred_by_type' in case and case['referred_by_id']:
            referred_by_id_obj = case['referred_by_id']
            # Convert to ObjectId if it's a string
            if isinstance(referred_by_id_obj, str):
                referred_by_id_obj = parse_object_id(referred_by_id_obj)
            if referred_by_id_obj:
                if case['referred_by_type'] == 'patient':
                    referred_by = db.patients.find_one({'_id': referred_by_id_obj})
                else:  # doctor
                    referred_by = db.doctors.find_one({'_id': referred_by_id_obj})
                if referred_by:
                    case['referred_by_name'] = referred_by.get('name', '')
        
        # Get case charges (patient charges)
        all_charges = list(db.case_charges.find({'case_id': parse_object_id(id)}).sort('created_at', -1))
        
        hospital_charges = []
        pathology_charges = []
        pharmacy_charges = []
        new_doctor_charges = []
        
        # Populate charge master names and doctor names
        for charge in all_charges:
            ctype = charge.get('charge_type', 'hospital')
            
            if 'charge_master_id' in charge:
                charge_master = db.charge_master.find_one({'_id': charge['charge_master_id']})
                if charge_master:
                    charge['charge_name'] = charge_master.get('name', '')
            
            # Populate doctor name if doctor_id exists
            if 'doctor_id' in charge:
                doctor = db.doctors.find_one({'_id': charge['doctor_id']})
                if doctor:
                    charge['doctor_name'] = doctor.get('name', '')
            
            # Split into categories
            if charge.get('is_doctor_charge'):
                # Normalize for doctor charges UI expectations
                charge['amount'] = charge.get('total_amount', 0)
                charge['date'] = charge.get('charge_date')
                new_doctor_charges.append(charge)
            elif ctype == 'pathology':
                pathology_charges.append(charge)
            elif ctype == 'pharmacy':
                pharmacy_charges.append(charge)
            else:
                hospital_charges.append(charge)
                
        case['charges'] = serialize_doc(hospital_charges)
        case['pathology_charges'] = serialize_doc(pathology_charges)
        case['pharmacy_charges'] = serialize_doc(pharmacy_charges)
        
        # Get case doctor charges (legacy)
        case_doctor_charges = list(db.case_doctor_charges.find({'case_id': parse_object_id(id)}))
        # Populate doctor names
        for charge in case_doctor_charges:
            if 'doctor_id' in charge:
                doctor = db.doctors.find_one({'_id': charge['doctor_id']})
                if doctor:
                    charge['doctor_name'] = doctor.get('name', '')
                    
        # Combine legacy and new doctor charges
        case['doctor_charges'] = serialize_doc(case_doctor_charges + new_doctor_charges)
        
        # Get appointments for this case
        case_appointments = list(db.appointments.find({'case_id': parse_object_id(id)}))
        # Populate patient names and doctor names
        for apt in case_appointments:
            if 'patient_id' in apt:
                patient = db.patients.find_one({'_id': apt['patient_id']})
                if patient:
                    apt['patient_name'] = patient.get('name', '')
            if 'doctor_id' in apt:
                doctor = db.doctors.find_one({'_id': apt['doctor_id']})
                if doctor:
                    apt['doctor_name'] = doctor.get('name', '')
        case['appointments'] = serialize_doc(case_appointments)
        
        # Get prescriptions for this case
        case_prescriptions = list(db.prescriptions.find({'case_id': parse_object_id(id)}).sort('created_at', -1))
        # Populate patient and doctor names
        for pres in case_prescriptions:
            if 'patient_id' in pres and pres['patient_id']:
                patient = db.patients.find_one({'_id': pres['patient_id']})
                if patient:
                    pres['patient_name'] = patient.get('name', '')
            if 'doctor_id' in pres and pres['doctor_id']:
                doctor = db.doctors.find_one({'_id': pres['doctor_id']})
                if doctor:
                    pres['doctor_name'] = doctor.get('name', '')
        case['prescriptions'] = serialize_doc(case_prescriptions)

        # Get case studies
        case_studies = list(db.case_studies.find({'case_id': parse_object_id(id)}).sort('created_at', -1))
        # Populate doctor names
        for study in case_studies:
            if 'doctor_id' in study and study['doctor_id']:
                doctor = db.doctors.find_one({'_id': study['doctor_id']})
                if doctor:
                    study['doctor_name'] = doctor.get('name', '')
        case['case_studies'] = serialize_doc(case_studies)
        
        return jsonify(serialize_doc(case))
    except Exception as e:
        logging.error(f"Error getting case: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/cases', methods=['POST'])
def create_case():
    try:
        data = request.get_json()
        
        # Convert patient_id to ObjectId if present
        if 'patient_id' in data and data['patient_id']:
            data['patient_id'] = parse_object_id(data['patient_id'])
        
        # Convert referred_by_id to ObjectId if present
        if 'referred_by_id' in data and data['referred_by_id']:
            data['referred_by_id'] = parse_object_id(data['referred_by_id'])
        
        # Generate case number
        if 'case_number' not in data or not data['case_number']:
            year = datetime.now().year
            counter_id = f'case_number_{year}'
            
            # Check if counter exists for this year
            counter = db.counters.find_one({'_id': counter_id})
            if not counter:
                # Initialize counter from existing cases if any, otherwise start at 999
                last_case = db.cases.find_one({'case_number': {'$regex': f'CASE-{year}-'}}, sort=[('case_number', -1)])
                if last_case:
                    try:
                        last_num = int(last_case['case_number'].split('-')[-1])
                        initial_seq = last_num
                    except (ValueError, IndexError):
                        initial_seq = 999
                else:
                    initial_seq = 999
                
                db.counters.update_one(
                    {'_id': counter_id},
                    {'$setOnInsert': {'seq': initial_seq}},
                    upsert=True
                )
            
            # Atomic increment
            counter = db.counters.find_one_and_update(
                {'_id': counter_id},
                {'$inc': {'seq': 1}},
                return_document=True
            )
            data['case_number'] = f'CASE-{year}-{counter["seq"]:04d}'
        
        # Parse admission_date if present
        if 'admission_date' in data and isinstance(data['admission_date'], str):
            try:
                data['admission_date'] = datetime.fromisoformat(data['admission_date'].replace('Z', '+00:00'))
            except ValueError:
                pass  # Keep as string if parsing fails, or handle error

        data['created_at'] = datetime.now()
        result = db.cases.insert_one(data)
        
        # Auto-add default IPD charges
        if data.get('case_type') == 'IPD':
            default_charges = list(db.charge_master.find({
                '$or': [
                    {'category': 'DEFAULT_CATEGORY_IPD'},
                    {'charge_category': 'DEFAULT_CATEGORY_IPD'}
                ]
            }))
            
            if default_charges:
                case_charges = []
                for charge in default_charges:
                    case_charges.append({
                        'case_id': result.inserted_id,
                        'charge_master_id': charge['_id'],
                        'charge_name': charge['name'],
                        'amount': charge.get('amount', 0),    # Stored as amount or unit_amount?
                        'unit_amount': charge.get('amount', 0),
                        'quantity': 1,
                        'total_amount': charge.get('amount', 0),
                        'charge_type': 'hospital', # Default type
                        'created_at': datetime.now(),
                        'updated_at': datetime.now()
                    })
                
                if case_charges:
                    db.case_charges.insert_many(case_charges)

        return jsonify({'id': str(result.inserted_id), 'message': 'Case created successfully'}), 201
    except Exception as e:
        logging.error(f"Error creating case: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/cases/<id>', methods=['PUT'])
def update_case(id):
    try:
        case_id_obj = parse_object_id(id)
        if is_case_closed(case_id_obj):
             # Allow admin to update closed cases
             logging.debug(f"Update Case Closed Check. Current Role: {session.get('role')}")
             if session.get('role') != 'admin':
                return jsonify({'error': 'Cannot update a closed case'}), 400
            
        data = request.get_json()
        
        # Check if trying to close the case
        if data.get('status') == 'closed':
            # Calculate total charges
            pipeline_charges = [
                {'$match': {'case_id': case_id_obj}},
                {'$group': {'_id': None, 'total': {'$sum': '$total_amount'}}}
            ]
            charges_result = list(db.case_charges.aggregate(pipeline_charges))
            total_charges = charges_result[0]['total'] if charges_result else 0
            
            # Add legacy doctor charges
            doctor_charges_legacy = list(db.case_doctor_charges.find({'case_id': case_id_obj}))
            total_charges += sum(c.get('amount', 0) for c in doctor_charges_legacy)
            
            # Calculate total payments
            pipeline_payments = [
                {'$match': {'case_id': case_id_obj}},
                {'$group': {'_id': None, 'total': {'$sum': '$amount'}}}
            ]
            payments_result = list(db.payments.aggregate(pipeline_payments))
            total_payments = payments_result[0]['total'] if payments_result else 0
            
            if total_payments < total_charges:
                return jsonify({'error': f'Cannot close case. Payment overdue. Total: {total_charges}, Paid: {total_payments}'}), 400

        # Convert patient_id to ObjectId if present
        if 'patient_id' in data and data['patient_id']:
            data['patient_id'] = parse_object_id(data['patient_id'])
        
        # Convert referred_by_id to ObjectId if present
        if 'referred_by_id' in data and data['referred_by_id']:
            data['referred_by_id'] = parse_object_id(data['referred_by_id'])
        
        data['updated_at'] = datetime.now()
        result = db.cases.update_one({'_id': case_id_obj}, {'$set': data})
        if result.modified_count:
            return jsonify({'message': 'Case updated successfully'})
        return jsonify({'message': 'Case updated successfully (No changes made)'}) # Handle no changes but valid request
    except Exception as e:
        logging.error(f"Error updating case: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/cases/<id>', methods=['DELETE'])
def delete_case(id):
    try:
        # Role check
        user_id = request.headers.get('X-User-Id')
        if not user_id:
            return jsonify({'error': 'Unauthorized'}), 401
        
        user = db.users.find_one({'_id': parse_object_id(user_id)})
        if not user or user.get('role') != 'admin':
            return jsonify({'error': 'Forbidden: Only admin can delete cases'}), 403

        if is_case_closed(id):
            return jsonify({'error': 'Cannot delete a closed case'}), 400

        result = db.cases.delete_one({'_id': parse_object_id(id)})
        if result.deleted_count:
            return jsonify({'message': 'Case deleted successfully'})
        return jsonify({'error': 'Case not found'}), 404
    except Exception as e:
        logging.error(f"Error deleting case: {e}")
        return jsonify({'error': str(e)}), 500

# ==================== APPOINTMENTS API ====================

@app.route('/api/appointments', methods=['GET'])
def get_appointments():
    try:
        page = int(request.args.get('page', 1))
        limit = int(request.args.get('limit', 10))
        skip = (page - 1) * limit
        
        # Build Query
        query = {}
        filters = []
        
        if request.args.get('patient_id'):
            filters.append({'patient_id': parse_object_id(request.args.get('patient_id'))})
            
        if request.args.get('case_id'):
            filters.append({'case_id': parse_object_id(request.args.get('case_id'))})

        if request.args.get('doctor_id'):
            filters.append({'doctor_id': parse_object_id(request.args.get('doctor_id'))})
            
        if request.args.get('date'):
            # Filter by exact date string match (simple approach) or range
            # Here assuming exact string match for 'YYYY-MM-DD'
            filters.append({'appointment_date': request.args.get('date')})
            
        if len(filters) > 1:
            query = {'$and': filters}
        elif len(filters) == 1:
            query = filters[0]
        
        # Get total count for pagination
        total = db.appointments.count_documents(query)
        
        # Get paginated results
        appointments = list(db.appointments.find(query).sort('created_at', -1).skip(skip).limit(limit))
        
        # Populate patient names and doctor names
        for apt in appointments:
            if 'patient_id' in apt:
                patient = db.patients.find_one({'_id': apt['patient_id']})
                if patient:
                    apt['patient_name'] = patient.get('name', '')
            if 'doctor_id' in apt:
                doctor = db.doctors.find_one({'_id': apt['doctor_id']})
                if doctor:
                    apt['doctor_name'] = doctor.get('name', '')
        
        return jsonify({
            'appointments': serialize_doc(appointments),
            'total': total,
            'page': page,
            'limit': limit
        })
    except Exception as e:
        logging.error(f"Error getting appointments: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/appointments', methods=['POST'])
def create_appointment():
    try:
        data = request.get_json()
        
        # Convert IDs to ObjectId if present
        if 'patient_id' in data and data['patient_id']:
            data['patient_id'] = parse_object_id(data['patient_id'])
        if 'case_id' in data and data['case_id']:
            case_id_obj = parse_object_id(data['case_id'])
            data['case_id'] = case_id_obj
            
            # Check if case is closed
            case = db.cases.find_one({'_id': case_id_obj})
            if case and case.get('status') == 'closed':
                return jsonify({'error': 'Cannot add appointments to a closed case'}), 400
        if 'doctor_id' in data and data['doctor_id']:
            data['doctor_id'] = parse_object_id(data['doctor_id'])
        
        data['created_at'] = datetime.now()
        result = db.appointments.insert_one(data)
        return jsonify({'id': str(result.inserted_id), 'message': 'Appointment created successfully'}), 201
    except Exception as e:
        logging.error(f"Error creating appointment: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/appointments/<id>', methods=['PUT'])
def update_appointment(id):
    try:
        appointment_id = parse_object_id(id)
        appointment = db.appointments.find_one({'_id': appointment_id})
        if appointment and is_case_closed(appointment.get('case_id')):
            return jsonify({'error': 'Cannot update appointment for a closed case'}), 400
            
        data = request.get_json()
        
        # Convert IDs to ObjectId if present
        if 'patient_id' in data and data['patient_id']:
            data['patient_id'] = parse_object_id(data['patient_id'])
        if 'case_id' in data and data['case_id']:
            data['case_id'] = parse_object_id(data['case_id'])
        if 'doctor_id' in data and data['doctor_id']:
            data['doctor_id'] = parse_object_id(data['doctor_id'])
        
        data['updated_at'] = datetime.now()
        result = db.appointments.update_one({'_id': parse_object_id(id)}, {'$set': data})
        if result.modified_count:
            return jsonify({'message': 'Appointment updated successfully'})
        return jsonify({'error': 'Appointment not found'}), 404
    except Exception as e:
        logging.error(f"Error updating appointment: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/appointments/<id>', methods=['DELETE'])
def delete_appointment(id):
    try:
        appointment_id = parse_object_id(id)
        appointment = db.appointments.find_one({'_id': appointment_id})
        if appointment and is_case_closed(appointment.get('case_id')):
            return jsonify({'error': 'Cannot delete appointment for a closed case'}), 400
            
        result = db.appointments.delete_one({'_id': appointment_id})
        if result.deleted_count:
            return jsonify({'message': 'Appointment deleted successfully'})
        return jsonify({'error': 'Appointment not found'}), 404
    except Exception as e:
        logging.error(f"Error deleting appointment: {e}")
        return jsonify({'error': str(e)}), 500

# ==================== PRESCRIPTIONS API ====================

@app.route('/api/prescriptions', methods=['GET'])
def get_prescriptions():
    try:
        page = int(request.args.get('page', 1))
        limit = int(request.args.get('limit', 10))
        skip = (page - 1) * limit
        case_id = request.args.get('case_id')
        
        # Build query
        query = {}
        if case_id:
            query['case_id'] = parse_object_id(case_id)
        
        # Get total count for pagination
        total = db.prescriptions.count_documents(query)
        
        # Get paginated results
        prescriptions = list(db.prescriptions.find(query).sort('created_at', -1).skip(skip).limit(limit))
        
        # Populate patient and doctor names
        for pres in prescriptions:
            if 'patient_id' in pres and pres['patient_id']:
                patient = db.patients.find_one({'_id': pres['patient_id']})
                if patient:
                    pres['patient_name'] = patient.get('name', '')
            if 'doctor_id' in pres and pres['doctor_id']:
                doctor = db.doctors.find_one({'_id': pres['doctor_id']})
                if doctor:
                    pres['doctor_name'] = doctor.get('name', '')
        
        return jsonify({
            'prescriptions': serialize_doc(prescriptions),
            'total': total,
            'page': page,
            'limit': limit
        })
    except Exception as e:
        logging.error(f"Error getting prescriptions: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/prescriptions', methods=['POST'])
def create_prescription():
    try:
        # Check if this is a file upload
        if 'file' in request.files:
            file = request.files['file']
            case_id = request.form.get('case_id')
            patient_id = request.form.get('patient_id')
            doctor_id = request.form.get('doctor_id')
            prescription_date = request.form.get('prescription_date')
            notes = request.form.get('notes', '')
            medications = request.form.get('medications', '')
            
            # Check if case is closed
            if case_id:
                case = db.cases.find_one({'_id': parse_object_id(case_id)})
                if case and case.get('status') == 'closed':
                    return jsonify({'error': 'Cannot add prescriptions to a closed case'}), 400
            
            if file and file.filename:
                # Generate unique filename
                filename = secure_filename(file.filename)
                file_ext = os.path.splitext(filename)[1]
                unique_filename = f"{uuid.uuid4()}{file_ext}"
                
                # Ensure upload directory exists
                upload_dir = os.path.join(app.root_path, 'static', 'uploads', 'prescriptions')
                os.makedirs(upload_dir, exist_ok=True)
                
                # Save file
                file_path = os.path.join(upload_dir, unique_filename)
                file.save(file_path)
                
                # Create prescription record
                prescription_data = {
                    'case_id': parse_object_id(case_id) if case_id else None,
                    'patient_id': parse_object_id(patient_id) if patient_id else None,
                    'doctor_id': parse_object_id(doctor_id) if doctor_id else None,
                    'prescription_date': datetime.strptime(prescription_date, '%Y-%m-%d') if prescription_date else datetime.now(),
                    'medications': medications,
                    'file_path': f'/static/uploads/prescriptions/{unique_filename}',
                    'file_name': filename,
                    'notes': notes,
                    'created_at': datetime.now()
                }
                
                result = db.prescriptions.insert_one(prescription_data)
                return jsonify({'id': str(result.inserted_id), 'message': 'Prescription uploaded successfully'}), 201
            else:
                return jsonify({'error': 'No file provided'}), 400
        else:
            # JSON data (for backward compatibility)
            data = request.get_json()
            # Convert IDs to ObjectId if present
            if 'case_id' in data and data['case_id']:
                case_id_obj = parse_object_id(data['case_id'])
                data['case_id'] = case_id_obj
                
                # Check if case is closed
                case = db.cases.find_one({'_id': case_id_obj})
                if case and case.get('status') == 'closed':
                    return jsonify({'error': 'Cannot add prescriptions to a closed case'}), 400
            if 'patient_id' in data and data['patient_id']:
                data['patient_id'] = parse_object_id(data['patient_id'])
            if 'doctor_id' in data and data['doctor_id']:
                data['doctor_id'] = parse_object_id(data['doctor_id'])
            if 'prescription_date' in data and data['prescription_date']:
                if isinstance(data['prescription_date'], str):
                    data['prescription_date'] = datetime.strptime(data['prescription_date'], '%Y-%m-%d')
            
            data['created_at'] = datetime.now()
            result = db.prescriptions.insert_one(data)
            return jsonify({'id': str(result.inserted_id), 'message': 'Prescription created successfully'}), 201
    except Exception as e:
        logging.error(f"Error creating prescription: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/prescriptions/<id>', methods=['PUT'])
def update_prescription(id):
    try:
        prescription_id = parse_object_id(id)
        prescription = db.prescriptions.find_one({'_id': prescription_id})
        if prescription and is_case_closed(prescription.get('case_id')):
            return jsonify({'error': 'Cannot update prescription for a closed case'}), 400
            
        data = request.get_json()
        data['updated_at'] = datetime.now()
        result = db.prescriptions.update_one({'_id': prescription_id}, {'$set': data})
        if result.modified_count:
            return jsonify({'message': 'Prescription updated successfully'})
        return jsonify({'error': 'Prescription not found'}), 404
    except Exception as e:
        logging.error(f"Error updating prescription: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/prescriptions/<id>', methods=['DELETE'])
def delete_prescription(id):
    try:
        prescription_id = parse_object_id(id)
        prescription = db.prescriptions.find_one({'_id': prescription_id})
        if prescription and is_case_closed(prescription.get('case_id')):
            return jsonify({'error': 'Cannot delete prescription for a closed case'}), 400
            
        result = db.prescriptions.delete_one({'_id': prescription_id})
        if result.deleted_count:
            return jsonify({'message': 'Prescription deleted successfully'})
        return jsonify({'error': 'Prescription not found'}), 404
    except Exception as e:
        logging.error(f"Error deleting prescription: {e}")
        return jsonify({'error': str(e)}), 500

# ==================== CHARGE MASTER API ====================

@app.route('/api/charge-master', methods=['GET'])
def get_charge_master():
    try:
        page = int(request.args.get('page', 1))
        limit = int(request.args.get('limit', 10))
        search = request.args.get('search', '').strip()
        sort_by = request.args.get('sortBy', 'created_at')
        sort_order = int(request.args.get('sortOrder', -1))
        skip = (page - 1) * limit
        
        # Build query for search
        query = {}
        if search:
            query = {
                '$or': [
                    {'name': {'$regex': search, '$options': 'i'}},
                    {'category': {'$regex': search, '$options': 'i'}},
                    {'charge_category': {'$regex': search, '$options': 'i'}}
                ]
            }
            
        # Get total count for pagination
        total = db.charge_master.count_documents(query)
        
        # Get paginated results
        charges = list(db.charge_master.find(query).sort(sort_by, sort_order).skip(skip).limit(limit))
        
        return jsonify({
            'charges': serialize_doc(charges),
            'total': total,
            'page': page,
            'limit': limit
        })
    except Exception as e:
        logging.error(f"Error getting charge master: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/charge-master/<id>', methods=['GET'])
def get_charge_master_item(id):
    try:
        charge = db.charge_master.find_one({'_id': parse_object_id(id)})
        if not charge:
            return jsonify({'error': 'Charge not found'}), 404
        return jsonify(serialize_doc(charge))
    except Exception as e:
        logging.error(f"Error getting charge: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/charge-master', methods=['POST'])
def create_charge_master():
    try:
        data = request.get_json()
        data['created_at'] = datetime.now()
        result = db.charge_master.insert_one(data)
        return jsonify({'id': str(result.inserted_id), 'message': 'Charge created successfully'}), 201
    except Exception as e:
        logging.error(f"Error creating charge: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/charge-master/<id>', methods=['PUT'])
def update_charge_master(id):
    try:
        data = request.get_json()
        data['updated_at'] = datetime.now()
        result = db.charge_master.update_one({'_id': parse_object_id(id)}, {'$set': data})
        if result.modified_count:
            return jsonify({'message': 'Charge updated successfully'})
        return jsonify({'error': 'Charge not found'}), 404
    except Exception as e:
        logging.error(f"Error updating charge: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/charge-master/<id>', methods=['DELETE'])
def delete_charge_master(id):
    try:
        result = db.charge_master.delete_one({'_id': parse_object_id(id)})
        if result.deleted_count:
            return jsonify({'message': 'Charge deleted successfully'})
        return jsonify({'error': 'Charge not found'}), 404
    except Exception as e:
        logging.error(f"Error deleting charge: {e}")
        return jsonify({'error': str(e)}), 500

# ==================== CHARGE CATEGORY MASTER API ====================

@app.route('/api/charge-category-master', methods=['GET'])
def get_charge_categories():
    try:
        categories = list(db.charge_category_master.find().sort('name', 1))
        return jsonify(serialize_doc(categories))
    except Exception as e:
        logging.error(f"Error getting charge categories: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/charge-category-master', methods=['POST'])
def create_charge_category():
    try:
        data = request.get_json()
        if not data.get('name'):
            return jsonify({'error': 'Category name is required'}), 400
            
        # Check if already exists (case-insensitive)
        name = data['name'].strip()
        existing = db.charge_category_master.find_one({'name': {'$regex': f'^{name}$', '$options': 'i'}})
        if existing:
            return jsonify(serialize_doc(existing)), 200
            
        data['name'] = name
        data['created_at'] = datetime.now()
        result = db.charge_category_master.insert_one(data)
        return jsonify({'id': str(result.inserted_id), 'message': 'Category created successfully'}), 201
    except Exception as e:
        logging.error(f"Error creating charge category: {e}")
        return jsonify({'error': str(e)}), 500

# ==================== CASE CHARGES API (Patient Charges) ====================

@app.route('/api/case-charges', methods=['GET'])
def get_case_charges():
    try:
        case_id = request.args.get('case_id')
        query = {}
        if case_id:
            query['case_id'] = parse_object_id(case_id)
        charges = list(db.case_charges.find(query))
        
        # Populate charge master names
        for charge in charges:
            if 'charge_master_id' in charge:
                charge_master = db.charge_master.find_one({'_id': charge['charge_master_id']})
                if charge_master:
                    charge['charge_name'] = charge_master.get('name', '')
        
        return jsonify(serialize_doc(charges))
    except Exception as e:
        logging.error(f"Error getting case charges: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/case-charges', methods=['POST'])
@login_required
def create_case_charge():
    input_type = 'json'
    if request.files:
        input_type = 'form'
        
    try:
        data = {}
        file_path = None
        if input_type == 'form':
            # Handle form data similar to cases
            data = request.form.to_dict()
            if 'file' in request.files:
                file = request.files['file']
                if file and file.filename:
                    unique_filename = f"{uuid.uuid4()}_{secure_filename(file.filename)}"
                    upload_path = os.path.join(app.config['UPLOAD_FOLDER'], unique_filename)
                    file.save(upload_path)
                    file_path = f"/static/uploads/{unique_filename}"
        else:
            data = request.get_json() or {}

        # Convert case_id and charge_master_id to ObjectId
        if 'case_id' in data and data['case_id']:
            case_id_obj = parse_object_id(data['case_id'])
            data['case_id'] = case_id_obj
            
            # Check if case is closed
            case = db.cases.find_one({'_id': case_id_obj})
            if case and case.get('status') == 'closed':
                return jsonify({'error': 'Cannot add charges to a closed case'}), 400
        else:
             return jsonify({'error': 'Case ID is required'}), 400

        if 'charge_master_id' in data and data['charge_master_id']:
            data['charge_master_id'] = parse_object_id(data['charge_master_id'])
        
        if 'doctor_id' in data and data['doctor_id']:
            data['doctor_id'] = parse_object_id(data['doctor_id'])
        
        # Determine charge type (default to 'hospital' if not provided)
        # Note: Frontend 'Add Doctor Charge' sends is_doctor_charge=True but maybe not charge_type
        if 'charge_type' not in data:
            data['charge_type'] = 'hospital'
            
        # Convert quantity and amounts to proper types
        if 'quantity' in data:
            data['quantity'] = int(data['quantity']) if data['quantity'] else 1
        else:
            data['quantity'] = 1
            
        if 'unit_amount' in data:
            data['unit_amount'] = float(data['unit_amount']) if data['unit_amount'] else 0.0
        
        if 'total_amount' in data:
            data['total_amount'] = float(data['total_amount']) if data['total_amount'] else 0.0
        elif 'quantity' in data and 'unit_amount' in data:
            # Calculate total_amount if not provided
            data['total_amount'] = float(data['quantity']) * float(data['unit_amount'])
        
        data['created_at'] = datetime.now()
        if 'charge_date' not in data:
            data['charge_date'] = datetime.now()
        else:
             # Ensure date format if string
             if isinstance(data['charge_date'], str):
                 try:
                     # Try parsing ISO format
                     data['charge_date'] = datetime.fromisoformat(data['charge_date'].replace('Z', '+00:00'))
                 except ValueError:
                     # Attempt strict parsing or fallback
                     pass

        if file_path:
            data['file_path'] = file_path

        result = db.case_charges.insert_one(data)
        
        # Sync Payout Logic (Auto-Create/Update)
        if 'doctor_id' in data and data['doctor_id']:
             sync_payout_for_case(data['case_id'], data['doctor_id'])
        
        return jsonify({'message': 'Case charge added successfully', 'id': str(result.inserted_id)}), 201
    except Exception as e:
        logging.error(f"Error creating case charge: {e}")
        return jsonify({'error': str(e)}), 500

def sync_payout_for_case(case_id, doctor_id):
    """
    Calculates total doctor share for a specific case and doctor.
    Creates or updates the Payout record.
    """
    try:
        # Calculate Total Doctor Share
        pipeline = [
            {
                '$match': {
                    'case_id': case_id,
                    'doctor_id': doctor_id
                }
            },
            {
                '$group': {
                    '_id': None,
                    'charges': {'$push': '$$ROOT'}
                }
            }
        ]
        
        agg = list(db.case_charges.aggregate(pipeline))
        if not agg:
            return 
            
        charges = agg[0]['charges']
        total_doc_amount = 0
        total_hospital_amount = 0
        
        for charge in charges:
            qty = charge.get('quantity', 1)
            total_hospital_amount += charge.get('total_amount', 0)
            
            doc_rate = 0
            # Check doctor specific rate
            dc = db.doctor_charges.find_one({
                'doctor_id': doctor_id,
                'charge_master_id': charge['charge_master_id']
            })
            
            if dc:
                doc_rate = dc.get('amount', 0)
            else:
                pass 
                
            total_doc_amount += doc_rate * qty

        # Find Existing Payout
        existing_payout = db.payouts.find_one({
            'case_id': case_id,
            'doctor_id': doctor_id
        })
        
        if existing_payout:
            update_data = {
                'total_charge_amount': total_hospital_amount,
                'doctor_charge_amount': total_doc_amount,
                'updated_at': datetime.now()
            }
            db.payouts.update_one({'_id': existing_payout['_id']}, {'$set': update_data})
        else:
            case = db.cases.find_one({'_id': case_id})
            doctor = db.doctors.find_one({'_id': doctor_id})
            
            new_payout = {
                'case_id': case_id,
                'doctor_id': doctor_id,
                'case_number': case.get('case_number') if case else '',
                'patient_name': case.get('patient_name') if case else '',
                'doctor_name': doctor.get('name') if doctor else '',
                'total_charge_amount': total_hospital_amount,
                'doctor_charge_amount': total_doc_amount,
                'payment_status': 'pending', 
                'created_at': datetime.now(),
                'updated_at': datetime.now()
            }
            if not new_payout['patient_name'] and case and case.get('patient_id'):
                 pat = db.patients.find_one({'_id': case['patient_id']})
                 if pat:
                     new_payout['patient_name'] = pat.get('name')
                     
            db.payouts.insert_one(new_payout)
            
    except Exception as e:
        logging.error(f"Error syncing payout for case {case_id}: {e}")

@app.route('/api/case-charges/<id>', methods=['PUT'])
@login_required
def update_case_charge(id):
    try:
        data = request.get_json()
        
        # Get the existing charge to find case_id
        existing_charge = db.case_charges.find_one({'_id': parse_object_id(id)})
        if not existing_charge:
            return jsonify({'error': 'Charge not found'}), 404
        
        case_id_obj = existing_charge.get('case_id')
        if case_id_obj:
            # Check if case is closed
            case = db.cases.find_one({'_id': case_id_obj})
            if case and case.get('status') == 'closed':
                return jsonify({'error': 'Cannot modify charges for a closed case'}), 400
        
        # Convert IDs to ObjectId if present
        if 'case_id' in data and data['case_id']:
            data['case_id'] = parse_object_id(data['case_id'])
        if 'charge_master_id' in data and data['charge_master_id']:
            data['charge_master_id'] = parse_object_id(data['charge_master_id'])
        if 'doctor_id' in data and data['doctor_id']:
            data['doctor_id'] = parse_object_id(data['doctor_id'])
        
        # Convert quantity and amounts to proper types
        if 'quantity' in data:
            data['quantity'] = int(data['quantity']) if data['quantity'] else 1
        if 'unit_amount' in data:
            data['unit_amount'] = float(data['unit_amount']) if data['unit_amount'] else 0.0
        if 'total_amount' in data:
            data['total_amount'] = float(data['total_amount']) if data['total_amount'] else 0.0
        elif 'quantity' in data and 'unit_amount' in data:
            # Calculate total_amount if not provided
            data['total_amount'] = float(data['quantity']) * float(data['unit_amount'])
            
        if 'charge_date' in data and isinstance(data['charge_date'], str):
             try:
                 data['charge_date'] = datetime.fromisoformat(data['charge_date'].replace('Z', '+00:00'))
             except:
                 pass
        
        data['updated_at'] = datetime.now()
        result = db.case_charges.update_one({'_id': parse_object_id(id)}, {'$set': data})
        if result.modified_count:
            return jsonify({'message': 'Case charge updated successfully'})
        return jsonify({'error': 'Case charge not found'}), 404
    except Exception as e:
        logging.error(f"Error updating case charge: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/case-charges/<id>', methods=['GET'])
def get_case_charge(id):
    try:
        charge = db.case_charges.find_one({'_id': parse_object_id(id)})
        if not charge:
            return jsonify({'error': 'Charge not found'}), 404
        return jsonify(serialize_doc(charge))
    except Exception as e:
        logging.error(f"Error getting case charge: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/case-charges/<id>', methods=['DELETE'])
@login_required
def delete_case_charge(id):
    try:
        # Role check handled by login_required (session based)
        # Verify user has permission if needed (admin or owns case?)
        # For now, just ensuring they are logged in.


        charge_id = parse_object_id(id)
        charge = db.case_charges.find_one({'_id': charge_id})
        if charge and is_case_closed(charge.get('case_id')):
            return jsonify({'error': 'Cannot delete charges for a closed case'}), 400
            
        result = db.case_charges.delete_one({'_id': charge_id})
        if result.deleted_count:
            return jsonify({'message': 'Case charge deleted successfully'})
        return jsonify({'error': 'Case charge not found'}), 404
    except Exception as e:
        logging.error(f"Error deleting case charge: {e}")
        return jsonify({'error': str(e)}), 500

# ==================== DOCTOR CHARGES API ====================

@app.route('/api/doctor-charges', methods=['GET'])
def get_doctor_charges():
    try:
        page = int(request.args.get('page', 1))
        limit = int(request.args.get('limit', 10))
        skip = (page - 1) * limit
        
        doctor_id = request.args.get('doctor_id')
        query = {}
        if doctor_id:
            query['doctor_id'] = parse_object_id(doctor_id)
        
        # Get total count for pagination
        total = db.doctor_charges.count_documents(query)
        
        # Get paginated results
        charges = list(db.doctor_charges.find(query).sort('created_at', -1).skip(skip).limit(limit))
        
        # Populate doctor names and charge master names
        for charge in charges:
            if 'doctor_id' in charge:
                doctor = db.doctors.find_one({'_id': charge['doctor_id']})
                if doctor:
                    charge['doctor_name'] = doctor.get('name', '')
            if 'charge_master_id' in charge:
                charge_master = db.charge_master.find_one({'_id': charge['charge_master_id']})
                if charge_master:
                    charge['charge_master_name'] = charge_master.get('name', '')
        
        return jsonify({
            'charges': serialize_doc(charges),
            'total': total,
            'page': page,
            'limit': limit
        })
    except Exception as e:
        logging.error(f"Error getting doctor charges: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/doctors-by-charge', methods=['GET'])
def get_doctors_by_charge():
    try:
        charge_master_id = request.args.get('charge_master_id')
        if not charge_master_id:
            return jsonify({'error': 'charge_master_id parameter is required'}), 400
        
        # Find all doctor_ids who have a specialized rate for this charge
        specialized_doctor_ids = [charge['doctor_id'] for charge in db.doctor_charges.find({'charge_master_id': parse_object_id(charge_master_id)}) if 'doctor_id' in charge]
        
        # Get all active doctors
        all_active_doctors = list(db.doctors.find({'isActive': {'$ne': False}}))
        
        doctors = []
        for doctor in all_active_doctors:
            doc_data = serialize_doc(doctor)
            # Add a flag if they have a specialized rate, might be useful for UI
            doc_data['has_specialized_rate'] = doctor['_id'] in specialized_doctor_ids
            doctors.append(doc_data)
        
        return jsonify(doctors)
    except Exception as e:
        logging.error(f"Error getting doctors by charge: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/doctor-charges/<id>', methods=['GET'])
def get_doctor_charge(id):
    try:
        charge = db.doctor_charges.find_one({'_id': parse_object_id(id)})
        if not charge:
            return jsonify({'error': 'Doctor charge not found'}), 404
        
        # Populate names
        if 'doctor_id' in charge:
            doctor = db.doctors.find_one({'_id': charge['doctor_id']})
            if doctor:
                charge['doctor_name'] = doctor.get('name', '')
        if 'charge_master_id' in charge:
            charge_master = db.charge_master.find_one({'_id': charge['charge_master_id']})
            if charge_master:
                charge['charge_master_name'] = charge_master.get('name', '')
        
        return jsonify(serialize_doc(charge))
    except Exception as e:
        logging.error(f"Error getting doctor charge: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/doctor-charges', methods=['POST'])
def create_doctor_charge():
    try:
        data = request.get_json()
        
        # Convert doctor_id and charge_master_id to ObjectId if present
        if 'doctor_id' in data and data['doctor_id']:
            doctor_id = parse_object_id(data['doctor_id'])
            data['doctor_id'] = doctor_id
            
            # Check if doctor is Inhouse
            doctor = db.doctors.find_one({'_id': doctor_id})
            if doctor and doctor.get('isInhouse'):
                return jsonify({'error': 'Cannot configure charges for Inhouse doctors'}), 400
                
        if 'charge_master_id' in data and data['charge_master_id']:
            data['charge_master_id'] = parse_object_id(data['charge_master_id'])
        
        data['created_at'] = datetime.now()
        result = db.doctor_charges.insert_one(data)
        return jsonify({'id': str(result.inserted_id), 'message': 'Doctor charge created successfully'}), 201
    except Exception as e:
        logging.error(f"Error creating doctor charge: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/doctor-charges/<id>', methods=['PUT'])
def update_doctor_charge(id):
    try:
        data = request.get_json()
        
        # Convert doctor_id and charge_master_id to ObjectId if present
        if 'doctor_id' in data and data['doctor_id']:
            data['doctor_id'] = parse_object_id(data['doctor_id'])
        if 'charge_master_id' in data and data['charge_master_id']:
            data['charge_master_id'] = parse_object_id(data['charge_master_id'])
        
        data['updated_at'] = datetime.now()
        result = db.doctor_charges.update_one({'_id': parse_object_id(id)}, {'$set': data})
        if result.modified_count:
            return jsonify({'message': 'Doctor charge updated successfully'})
        return jsonify({'error': 'Doctor charge not found'}), 404
    except Exception as e:
        logging.error(f"Error updating doctor charge: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/doctor-charges/<id>', methods=['DELETE'])
def delete_doctor_charge(id):
    try:
        result = db.doctor_charges.delete_one({'_id': parse_object_id(id)})
        if result.deleted_count:
            return jsonify({'message': 'Doctor charge deleted successfully'})
        return jsonify({'error': 'Doctor charge not found'}), 404
    except Exception as e:
        logging.error(f"Error deleting doctor charge: {e}")
        return jsonify({'error': str(e)}), 500

# ==================== CASE DOCTOR CHARGES API ====================

@app.route('/api/case-doctor-charges', methods=['GET'])
def get_case_doctor_charges():
    try:
        case_id = request.args.get('case_id')
        query = {}
        if case_id:
            query['case_id'] = parse_object_id(case_id)
        
        charges = list(db.case_doctor_charges.find(query))
        
        # Populate doctor names
        for charge in charges:
            if 'doctor_id' in charge:
                doctor = db.doctors.find_one({'_id': charge['doctor_id']})
                if doctor:
                    charge['doctor_name'] = doctor.get('name', '')
        
        return jsonify(serialize_doc(charges))
    except Exception as e:
        logging.error(f"Error getting case doctor charges: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/case-doctor-charges/<id>', methods=['GET'])
def get_case_doctor_charge(id):
    try:
        # Try finding in case_charges first
        charge = db.case_charges.find_one({'_id': parse_object_id(id)})
        if not charge:
            # Fallback to legacy collection
            charge = db.case_doctor_charges.find_one({'_id': parse_object_id(id)})
            
        if not charge:
            return jsonify({'error': 'Case doctor charge not found'}), 404
        
        # Populate doctor name
        if 'doctor_id' in charge:
            doctor = db.doctors.find_one({'_id': charge['doctor_id']})
            if doctor:
                charge['doctor_name'] = doctor.get('name', '')
        
        return jsonify(serialize_doc(charge))
    except Exception as e:
        logging.error(f"Error getting case doctor charge: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/case-doctor-charges', methods=['POST'])
def create_case_doctor_charge():
    try:
        data = request.get_json()
        
        # Parse fields for case_charges schema
        case_id = data.get('case_id')
        doctor_id = data.get('doctor_id')
        charge_master_id = data.get('charge_master_id')
        amount = float(data.get('amount', 0))
        date = data.get('charge_date')
        date = data.get('charge_date')
        notes = data.get('notes', '')
        
        if is_case_closed(case_id):
            return jsonify({'error': 'Cannot add doctor charges to a closed case'}), 400
        
        charge_name = 'Doctor Charge'
        charge_type = 'Consultation'
        
        if charge_master_id:
            cm = db.charge_master.find_one({'_id': parse_object_id(charge_master_id)})
            if cm:
                charge_name = cm.get('name', 'Doctor Charge')
                charge_type = cm.get('category', 'Consultation')

        # Create record for case_charges
        charge_record = {
            'case_id': parse_object_id(case_id) if case_id else None,
            'doctor_id': parse_object_id(doctor_id) if doctor_id else None,
            'charge_master_id': parse_object_id(charge_master_id) if charge_master_id else None,
            'charge_name': charge_name,
            'charge_type': charge_type,
            'quantity': 1,
            'rate': amount,
            'total_amount': amount,
            'charge_date': date if date else datetime.now(),
            'notes': notes,
            'created_at': datetime.now(),
            'is_doctor_charge': True # Flag to distinguish if needed
        }
        
        result = db.case_charges.insert_one(charge_record)
        return jsonify({'id': str(result.inserted_id), 'message': 'Doctor charge added successfully'}), 201
    except Exception as e:
        logging.error(f"Error creating doctor charge: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/case-doctor-charges/<id>', methods=['PUT'])
def update_case_doctor_charge(id):
    try:
        data = request.get_json()
        
        # Check if it's in case_charges or case_doctor_charges
        target_id = parse_object_id(id)
        is_new_collection = db.case_charges.find_one({'_id': target_id}) is not None
        is_new_collection = db.case_charges.find_one({'_id': target_id}) is not None
        collection = db.case_charges if is_new_collection else db.case_doctor_charges
        
        existing = collection.find_one({'_id': target_id})
        if existing and is_case_closed(existing.get('case_id')):
             return jsonify({'error': 'Cannot modify doctor charges for a closed case'}), 400
        
        # Fields to update
        update_fields = {}
        if 'doctor_id' in data: update_fields['doctor_id'] = parse_object_id(data['doctor_id'])
        if 'charge_master_id' in data: 
            cm_id = data['charge_master_id']
            update_fields['charge_master_id'] = parse_object_id(cm_id)
            cm = db.charge_master.find_one({'_id': parse_object_id(cm_id)})
            if cm:
                update_fields['charge_name'] = cm.get('name', 'Doctor Charge')
                update_fields['charge_type'] = cm.get('category', 'Consultation')
        
        if 'amount' in data:
            if is_new_collection:
                update_fields['rate'] = float(data['amount'])
                update_fields['total_amount'] = float(data['amount'])
            else:
                update_fields['amount'] = float(data['amount'])
        
        if 'charge_date' in data: update_fields['charge_date'] = data['charge_date']
        if 'notes' in data: update_fields['notes'] = data['notes']
        update_fields['updated_at'] = datetime.now()
        
        collection.update_one({'_id': target_id}, {'$set': update_fields})
        return jsonify({'message': 'Doctor charge updated successfully'})
    except Exception as e:
        logging.error(f"Error updating case doctor charge: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/case-doctor-charges/<id>', methods=['DELETE'])
def delete_case_doctor_charge(id):
    try:
        charge_id = parse_object_id(id)
        
        # Check if closed (Checking both collections)
        existing = db.case_charges.find_one({'_id': charge_id}) or db.case_doctor_charges.find_one({'_id': charge_id})
        if existing and is_case_closed(existing.get('case_id')):
            return jsonify({'error': 'Cannot delete doctor charges for a closed case'}), 400

        # Try deleting from case_charges first (new schema)
        result = db.case_charges.delete_one({'_id': charge_id})
        if result.deleted_count:
             return jsonify({'message': 'Case doctor charge deleted successfully'})
             
        # Fallback to legacy
        result = db.case_doctor_charges.delete_one({'_id': charge_id})
        if result.deleted_count:
            return jsonify({'message': 'Case doctor charge deleted successfully'})
            
        return jsonify({'error': 'Case doctor charge not found'}), 404
    except Exception as e:
        logging.error(f"Error deleting case doctor charge: {e}")
        return jsonify({'error': str(e)}), 500

# ==================== BILLS API (Optimized with Aggregation) ====================

@app.route('/api/bills', methods=['GET'])
def get_bills():
    try:
        case_id = request.args.get('case_id')
        query = {}
        if case_id:
            query['case_id'] = parse_object_id(case_id)
        
        # Use aggregation pipeline for efficient joins
        pipeline = [
            {'$match': query} if query else {'$match': {}},
            {
                '$lookup': {
                    'from': 'cases',
                    'localField': 'case_id',
                    'foreignField': '_id',
                    'as': 'case'
                }
            },
            {'$unwind': {'path': '$case', 'preserveNullAndEmptyArrays': True}},
            {
                '$lookup': {
                    'from': 'patients',
                    'localField': 'case.patient_id',
                    'foreignField': '_id',
                    'as': 'patient'
                }
            },
            {'$unwind': {'path': '$patient', 'preserveNullAndEmptyArrays': True}},
            {
                '$project': {
                    '_id': 1,
                    'case_id': 1,
                    'total_amount': 1,
                    'paid_amount': 1,
                    'balance_amount': 1,
                    'status': 1,
                    'created_at': 1,
                    'case_number': '$case.case_number',
                    'patient_name': '$patient.name'
                }
            }
        ]
        
        bills = list(db.bills.aggregate(pipeline))
        return jsonify(serialize_doc(bills))
    except Exception as e:
        logging.error(f"Error getting bills: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/bills', methods=['POST'])
def create_bill():
    try:
        data = request.get_json()
        
        # Calculate bill amounts from case charges
        case_id = parse_object_id(data.get('case_id'))
        case_charges = list(db.case_charges.find({'case_id': case_id}))
        total_amount = sum(charge.get('total_amount', 0) for charge in case_charges)
        
        data['total_amount'] = data.get('total_amount', total_amount)
        data['paid_amount'] = data.get('paid_amount', 0)
        data['balance_amount'] = data['total_amount'] - data['paid_amount']
        data['status'] = data.get('status', 'pending')
        data['created_at'] = datetime.now()
        
        result = db.bills.insert_one(data)
        return jsonify({'id': str(result.inserted_id), 'message': 'Bill created successfully'}), 201
    except Exception as e:
        logging.error(f"Error creating bill: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/bills/<id>', methods=['PUT'])
def update_bill(id):
    try:
        data = request.get_json()
        data['updated_at'] = datetime.now()
        
        # Recalculate balance if amounts changed
        if 'paid_amount' in data or 'total_amount' in data:
            bill = db.bills.find_one({'_id': parse_object_id(id)})
            if bill:
                total = data.get('total_amount', bill.get('total_amount', 0))
                paid = data.get('paid_amount', bill.get('paid_amount', 0))
                data['balance_amount'] = total - paid
        
        result = db.bills.update_one({'_id': parse_object_id(id)}, {'$set': data})
        if result.modified_count:
            return jsonify({'message': 'Bill updated successfully'})
        return jsonify({'error': 'Bill not found'}), 404
    except Exception as e:
        logging.error(f"Error updating bill: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/bills/<id>', methods=['DELETE'])
def delete_bill(id):
    try:
        result = db.bills.delete_one({'_id': parse_object_id(id)})
        if result.deleted_count:
            return jsonify({'message': 'Bill deleted successfully'})
        return jsonify({'error': 'Bill not found'}), 404
    except Exception as e:
        logging.error(f"Error deleting bill: {e}")
        return jsonify({'error': str(e)}), 500

# ==================== PAYMENTS API (Optimized with Aggregation) ====================

@app.route('/api/payments', methods=['GET'])
def get_payments():
    try:
        case_id = request.args.get('case_id')
        query = {}
        if case_id:
            query['case_id'] = parse_object_id(case_id)
        
        # Use aggregation pipeline for efficient joins
        pipeline = [
            {'$match': query} if query else {'$match': {}},
            {
                '$lookup': {
                    'from': 'cases',
                    'localField': 'case_id',
                    'foreignField': '_id',
                    'as': 'case'
                }
            },
            {'$unwind': {'path': '$case', 'preserveNullAndEmptyArrays': True}},
            {
                '$lookup': {
                    'from': 'patients',
                    'localField': 'case.patient_id',
                    'foreignField': '_id',
                    'as': 'patient'
                }
            },
            {'$unwind': {'path': '$patient', 'preserveNullAndEmptyArrays': True}},
            {
                '$project': {
                    '_id': 1,
                    'case_id': 1,
                    'amount': 1,
                    'payment_mode': 1,
                    'payment_date': 1,
                    'notes': 1,
                    'created_at': 1,
                    'case_number': '$case.case_number',
                    'patient_name': '$patient.name'
                }
            }
        ]
        
        payments = list(db.payments.aggregate(pipeline))
        return jsonify(serialize_doc(payments))
    except Exception as e:
        logging.error(f"Error getting payments: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/payments', methods=['POST'])
def create_payment():
    try:
        data = request.get_json()
        if not data:
            return jsonify({'error': 'No data provided'}), 400
            
        case_id = data.get('case_id')
        if case_id and is_case_closed(case_id):
            return jsonify({'error': 'Cannot record payments for a closed case'}), 400
            
        data['created_at'] = datetime.now()
        
        # Ensure case_id and patient_id are ObjectIds
        case_id = parse_object_id(data.get('case_id'))
        if case_id:
            data['case_id'] = case_id
            
            # Get patient_id from case if not provided
            if 'patient_id' not in data or not data.get('patient_id'):
                case = db.cases.find_one({'_id': case_id})
                if case and 'patient_id' in case:
                    data['patient_id'] = parse_object_id(case['patient_id'])
            else:
                data['patient_id'] = parse_object_id(data['patient_id'])
            
            # Update bill paid amount if bill exists (optional - don't fail if bill doesn't exist)
            bill = db.bills.find_one({'case_id': case_id})
            if bill:
                new_paid = bill.get('paid_amount', 0) + data.get('amount', 0)
                db.bills.update_one(
                    {'_id': bill['_id']},
                    {
                        '$set': {
                            'paid_amount': new_paid,
                            'balance_amount': bill.get('total_amount', 0) - new_paid,
                            'status': 'paid' if new_paid >= bill.get('total_amount', 0) else 'partial'
                        }
                    }
                )
        elif 'patient_id' in data:
            data['patient_id'] = parse_object_id(data['patient_id'])
        
        # Ensure amount is a float
        if 'amount' in data:
            data['amount'] = float(data['amount']) if data['amount'] else 0.0
        
        # Ensure payment_date is a datetime if provided
        if 'payment_date' in data and data['payment_date']:
            if isinstance(data['payment_date'], str):
                try:
                    # Try ISO format first
                    data['payment_date'] = datetime.fromisoformat(data['payment_date'].replace('Z', '+00:00'))
                except:
                    try:
                        # Try YYYY-MM-DD format
                        data['payment_date'] = datetime.strptime(data['payment_date'], '%Y-%m-%d')
                    except:
                        try:
                            # Try parsing as ISO string
                            data['payment_date'] = datetime.fromisoformat(data['payment_date'])
                        except:
                            # If all fail, use current date
                            data['payment_date'] = datetime.now()
                            logging.warning(f"Could not parse payment_date: {data.get('payment_date')}, using current date")
        else:
            # If no payment_date provided, use current date
            data['payment_date'] = datetime.now()
        
        # Log the data being saved
        logging.info(f"Creating payment with data: case_id={case_id}, amount={data.get('amount')}, payment_mode={data.get('payment_mode')}")
        
        # Save payment to payments collection (NO payout updates - payments are independent)
        result = db.payments.insert_one(data)
        payment_id = str(result.inserted_id)
        logging.info(f"Payment created successfully: {payment_id} for case: {case_id}, amount: {data.get('amount')}")
        return jsonify({'id': payment_id, 'message': 'Payment created successfully'}), 201
    except Exception as e:
        logging.error(f"Error creating payment: {e}", exc_info=True)
        return jsonify({'error': f'Failed to create payment: {str(e)}'}), 500

@app.route('/api/payments/<id>', methods=['PUT'])
def update_payment(id):
    try:
        payment_id = parse_object_id(id)
        payment = db.payments.find_one({'_id': payment_id})
        if payment and is_case_closed(payment.get('case_id')):
            return jsonify({'error': 'Cannot update payment for a closed case'}), 400
            
        data = request.get_json()
        data['updated_at'] = datetime.now()
        result = db.payments.update_one({'_id': payment_id}, {'$set': data})
        if result.modified_count:
            return jsonify({'message': 'Payment updated successfully'})
        return jsonify({'error': 'Payment not found'}), 404
    except Exception as e:
        logging.error(f"Error updating payment: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/payments/<id>', methods=['DELETE'])
def delete_payment(id):
    try:
        payment_id = parse_object_id(id)
        payment = db.payments.find_one({'_id': payment_id})
        if payment and is_case_closed(payment.get('case_id')):
            return jsonify({'error': 'Cannot delete payment for a closed case'}), 400
            
        result = db.payments.delete_one({'_id': payment_id})
        if result.deleted_count:
            return jsonify({'message': 'Payment deleted successfully'})
        return jsonify({'error': 'Payment not found'}), 404
    except Exception as e:
        logging.error(f"Error deleting payment: {e}")
        return jsonify({'error': str(e)}), 500

# ==================== DOCTOR PAYOUTS API ====================

@app.route('/api/doctor-payouts', methods=['GET'])
def get_doctor_payouts():
    try:
        date = request.args.get('date')  # Format: YYYY-MM-DD
        if not date:
            return jsonify({'error': 'Date parameter is required'}), 400
        
        # Parse date
        date_obj = datetime.strptime(date, '%Y-%m-%d')
        start_date = datetime.combine(date_obj.date(), datetime.min.time())
        end_date = datetime.combine(date_obj.date(), datetime.max.time())
        
        # Find all cases for the date (using admission_date)
        cases = list(db.cases.find({
            'admission_date': {
                '$gte': start_date,
                '$lte': end_date
            }
        }))
        
        payouts = []
        for case in cases:
            # Get patient name
            patient_name = ''
            if 'patient_id' in case:
                patient = db.patients.find_one({'_id': case['patient_id']})
                if patient:
                    patient_name = patient.get('name', '')
            
            # Calculate total charge amount (from case_charges)
            case_charges = list(db.case_charges.find({'case_id': case['_id']}))
            total_charge_amount = sum(charge.get('total_amount', 0) for charge in case_charges)
            
            # Get doctor charge amounts from doctor_charges collection based on charge_master_id
            # For each case_charge with doctor_id, look up the doctor_charges amount for that category
            doctor_charges_by_doctor = {}
            for charge in case_charges:
                if 'doctor_id' in charge and charge['doctor_id'] and 'charge_master_id' in charge and charge['charge_master_id']:
                    doctor_id = charge['doctor_id']
                    charge_master_id = charge['charge_master_id']
                    quantity = charge.get('quantity', 1)
                    
                    # Look up doctor_charges to get the amount for this doctor and charge category
                    doctor_charge = db.doctor_charges.find_one({
                        'doctor_id': doctor_id,
                        'charge_master_id': charge_master_id
                    })
                    
                    if doctor_charge:
                        # Get the amount from doctor_charges and multiply by quantity
                        doctor_charge_amount = doctor_charge.get('amount', 0) * quantity
                        
                        doctor_id_str = str(doctor_id)
                        if doctor_id_str not in doctor_charges_by_doctor:
                            doctor_charges_by_doctor[doctor_id_str] = {
                                'doctor_id': doctor_id_str,
                                'doctor_name': '',
                                'total_amount': 0
                            }
                        doctor_charges_by_doctor[doctor_id_str]['total_amount'] += doctor_charge_amount
            
            # Populate doctor names
            for doctor_id_str, doctor_charge_data in doctor_charges_by_doctor.items():
                doctor = db.doctors.find_one({'_id': parse_object_id(doctor_id_str)})
                if doctor:
                    doctor_charge_data['doctor_name'] = doctor.get('name', '')
            
            # If no doctor charges found in case_charges, check case_doctor_charges for backward compatibility
            if not doctor_charges_by_doctor and 'doctor_id' in case:
                doctor = db.doctors.find_one({'_id': case['doctor_id']})
                if doctor:
                    case_doc_charges = list(db.case_doctor_charges.find({'case_id': case['_id'], 'doctor_id': case['doctor_id']}))
                    if case_doc_charges:
                        total_doc_amount = sum(charge.get('amount', 0) for charge in case_doc_charges)
                        doctor_charges_by_doctor[str(case['doctor_id'])] = {
                            'doctor_id': str(case['doctor_id']),
                            'doctor_name': doctor.get('name', ''),
                            'total_amount': total_doc_amount
                        }
            
            # Calculate total doctor charge amount for the case
            total_doctor_charge_amount = sum(doc_charge['total_amount'] for doc_charge in doctor_charges_by_doctor.values())
            
            payouts.append({
                'case_id': str(case['_id']),
                'case_number': case.get('case_number', ''),
                'patient_name': patient_name,
                'case_type': case.get('case_type', ''),
                'total_charge_amount': total_charge_amount,
                'doctor_charge_amount': total_doctor_charge_amount,
                'doctor_charges': list(doctor_charges_by_doctor.values())  # List of doctor charges with doctor names
            })
        
        return jsonify(serialize_doc(payouts))
    except Exception as e:
        logging.error(f"Error getting doctor payouts: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/cases/<id>/payout-summary', methods=['GET'])
def get_case_payout_summary(id):
    try:
        case_id = parse_object_id(id)
        case = db.cases.find_one({'_id': case_id})
        if not case:
            return jsonify({'error': 'Case not found'}), 404
            
        # Calculate totals
        total_charge_amount = 0
        doctor_charge_amount = 0
        details = []
        
        # 1. Process case_charges
        case_charges = list(db.case_charges.find({'case_id': case_id}))
        for cc in case_charges:
            qty = cc.get('quantity', 1)
            # Hospital amount
            total_charge_amount += cc.get('total_amount', 0)
            
            # Doctor amount
            doc_amount = 0
            charge_name = 'Unknown Charge'
            
            if cc.get('charge_master_id'):
                cm = db.charge_master.find_one({'_id': cc.get('charge_master_id')})
                if cm:
                    charge_name = cm.get('name', 'Unknown')
                
                # Check for specific doctor charge config
                if cc.get('doctor_id'):
                    dc = db.doctor_charges.find_one({
                        'doctor_id': cc['doctor_id'],
                        'charge_master_id': cc['charge_master_id']
                    })
                    if dc:
                        doc_amount = dc.get('amount', 0) * qty
            
            doctor_charge_amount += doc_amount
            if doc_amount > 0:
                details.append({
                    'name': charge_name,
                    'amount': doc_amount,
                    'quantity': qty
                })

        # 2. Process case_doctor_charges (legacy/fallback)
        if not case_charges:
            case_doc_charges = list(db.case_doctor_charges.find({'case_id': case_id}))
            for cdc in case_doc_charges:
                amount = cdc.get('amount', 0)
                doctor_charge_amount += amount
                details.append({
                    'name': cdc.get('charge_name', 'Legacy Charge'),
                    'amount': amount,
                    'quantity': 1
                })
                # Note: Legacy charges might not have contributed to total_charge_amount in the same way, 
                # but usually they are added to case total. Checking case_charges is safer.
        
        return jsonify({
            'total_charge_amount': total_charge_amount,
            'doctor_charge_amount': doctor_charge_amount,
            'details': details
        })
    except Exception as e:
        logging.error(f"Error getting payout summary: {e}")
        return jsonify({'error': str(e)}), 500

# ==================== PAYOUTS API ====================

@app.route('/api/payouts', methods=['GET'])
def get_payouts():
    try:
        page = int(request.args.get('page', 1))
        limit = int(request.args.get('limit', 10))
        skip = (page - 1) * limit
        
        # Optional filters
        case_id = request.args.get('case_id')
        doctor_id = request.args.get('doctor_id')
        payment_status = request.args.get('payment_status')
        date = request.args.get('date')  # Format: YYYY-MM-DD (single date)
        start_date = request.args.get('start_date')  # Format: YYYY-MM-DD
        end_date = request.args.get('end_date')  # Format: YYYY-MM-DD
        
        query = {}
        if case_id:
            query['case_id'] = parse_object_id(case_id)
        if doctor_id:
            query['doctor_id'] = parse_object_id(doctor_id)
        if payment_status:
            query['payment_status'] = payment_status
        
        # Date range filtering (priority: start_date/end_date > date)
        if start_date and end_date:
            start_date_obj = datetime.strptime(start_date, '%Y-%m-%d')
            end_date_obj = datetime.strptime(end_date, '%Y-%m-%d')
            query['date_time'] = {
                '$gte': datetime.combine(start_date_obj.date(), datetime.min.time()),
                '$lte': datetime.combine(end_date_obj.date(), datetime.max.time())
            }
        elif date:
            date_obj = datetime.strptime(date, '%Y-%m-%d')
            query['date_time'] = {
                '$gte': datetime.combine(date_obj.date(), datetime.min.time()),
                '$lte': datetime.combine(date_obj.date(), datetime.max.time())
            }
        
        # Get total count
        total = db.payouts.count_documents(query)
        
        # Get paginated results
        payouts = list(db.payouts.find(query).sort('date_time', -1).skip(skip).limit(limit))
        
        # Populate doctor names
        # Populate doctor names and charge details
        for payout in payouts:
            if 'doctor_id' in payout and payout['doctor_id']:
                doctor = db.doctors.find_one({'_id': payout['doctor_id']})
                if doctor:
                    payout['doctor_name'] = doctor.get('name', '')
            
            # Fetch charge breakdown
            if 'case_id' in payout and 'doctor_id' in payout:
                charge_details = []
                case_charges = list(db.case_charges.find({
                    'case_id': payout['case_id'], 
                    'doctor_id': payout['doctor_id']
                }))
                
                for cc in case_charges:
                    cm = db.charge_master.find_one({'_id': cc.get('charge_master_id')})
                    charge_name = cm.get('name', 'Unknown') if cm else 'Unknown'
                    
                    # Calculate amount for doctor (qty * rate from doctor_charges)
                    # Note: case_charges stores the HOSPITAL rate. We need doctor rate.
                    doc_rate = 0
                    if cc.get('doctor_id') and cc.get('charge_master_id'):
                        dc = db.doctor_charges.find_one({
                            'doctor_id': cc['doctor_id'],
                            'charge_master_id': cc['charge_master_id']
                        })
                        if dc:
                            doc_rate = dc.get('amount', 0)
                    
                    qty = cc.get('quantity', 1)
                    total = doc_rate * qty
                    
                    charge_details.append({
                        'name': charge_name,
                        'amount': total,
                        'quantity': qty
                    })
                
                # Fallback to case_doctor_charges if no case_charges found (backward compatibility)
                if not charge_details:
                    cdc = list(db.case_doctor_charges.find({
                        'case_id': payout['case_id'],
                        'doctor_id': payout['doctor_id']
                    }))
                    for c in cdc:
                        charge_details.append({
                            'name': c.get('charge_name', 'Charge'),
                            'amount': c.get('amount', 0),
                            'quantity': 1
                        })
                        
                payout['charge_details'] = charge_details
        
        return jsonify({
            'payouts': serialize_doc(payouts),
            'total': total,
            'page': page,
            'limit': limit
        })
    except Exception as e:
        logging.error(f"Error getting payouts: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/payouts', methods=['POST'])
def create_payout():
    try:
        data = request.get_json()
        
        # Convert IDs to ObjectId if present
        if 'case_id' in data and data['case_id']:
            data['case_id'] = parse_object_id(data['case_id'])
        if 'doctor_id' in data and data['doctor_id']:
            data['doctor_id'] = parse_object_id(data['doctor_id'])
        
        # Convert amounts to proper types
        if 'total_charge_amount' in data:
            data['total_charge_amount'] = float(data['total_charge_amount']) if data['total_charge_amount'] else 0.0
        if 'doctor_charge_amount' in data:
            data['doctor_charge_amount'] = float(data['doctor_charge_amount']) if data['doctor_charge_amount'] else 0.0
        if 'partial_payment_amount' in data:
            data['partial_payment_amount'] = float(data['partial_payment_amount']) if data['partial_payment_amount'] else 0.0
        
        # Set default values
        if 'payment_status' not in data:
            data['payment_status'] = 'pending'
        if 'date_time' not in data:
            data['date_time'] = datetime.now()
        
        data['created_at'] = datetime.now()
        result = db.payouts.insert_one(data)
        return jsonify({'id': str(result.inserted_id), 'message': 'Payout created successfully'}), 201
    except Exception as e:
        logging.error(f"Error creating payout: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/payouts/pending', methods=['GET'])
def get_pending_payouts():
    try:
        # Find cases with doctor charges
        # Method 1: Aggregate case_charges to find unique (case_id, doctor_id)
        pipeline = [
            {
                '$match': {
                    'doctor_id': {'$exists': True, '$ne': None},
                    'charge_master_id': {'$exists': True}
                }
            },
            {
                '$group': {
                    '_id': {
                        'case_id': '$case_id',
                        'doctor_id': '$doctor_id'
                    },
                    'charges': {'$push': '$$ROOT'}
                }
            }
        ]
        
        pending_payouts = []
        grouped_charges = list(db.case_charges.aggregate(pipeline))
        
        for group in grouped_charges:
            case_id = group['_id']['case_id']
            doctor_id = group['_id']['doctor_id']
            
            # Check if payout already exists for this case+doctor
            existing_payout = db.payouts.find_one({
                'case_id': case_id,
                'doctor_id': doctor_id,
                'payment_status': {'$ne': 'cancelled'} 
            })
            
            if existing_payout:
                continue
                
            # Get Context
            case = db.cases.find_one({'_id': case_id})
            doctor = db.doctors.find_one({'_id': doctor_id})
            
            if not case or not doctor:
                continue
                
            # Calculate Amount
            total_doc_amount = 0
            try:
                for charge in group['charges']:
                    qty = charge.get('quantity', 1)
                    doc_rate = 0
                    
                    # Check doctor specific rate
                    dc = db.doctor_charges.find_one({
                        'doctor_id': doctor_id,
                        'charge_master_id': charge['charge_master_id']
                    })
                    
                    if dc:
                        doc_rate = dc.get('amount', 0)
                    else:
                        # Fallback logic here if needed
                        pass 
                        
                    total_doc_amount += doc_rate * qty
                
                if total_doc_amount > 0:
                    created_at = case.get('created_at')
                    date_str = None
                    if isinstance(created_at, datetime):
                        date_str = created_at.strftime('%Y-%m-%d')
                    elif isinstance(created_at, str):
                        date_str = created_at
                    
                    patient_name = case.get('patient_name')
                    if not patient_name and case.get('patient_id'):
                        patient = db.patients.find_one({'_id': case.get('patient_id')})
                        if patient:
                            patient_name = patient.get('name')

                    pending_payouts.append({
                        'case_id': str(case_id),
                        'doctor_id': str(doctor_id),
                        'case_number': case.get('case_number'),
                        'patient_name': patient_name,
                        'doctor_name': doctor.get('name'),
                        'amount': total_doc_amount,
                        'date': date_str
                    })
            except Exception as loop_e:
                logging.error(f"Error processing pending payout for case {case_id}: {loop_e}")
                continue
                
        return jsonify(pending_payouts)
    except Exception as e:
        logging.error(f"Error getting pending payouts: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/payouts/<id>', methods=['PUT'])
def update_payout(id):
    try:
        data = request.get_json()
        update_data = {'updated_at': datetime.now()}
        
        # Allow updating payment_status, payment_comment, payment_mode, payment_reference_number, payment_date, partial_payment_amount
        allowed_fields = ['payment_status', 'payment_comment', 'payment_mode', 'payment_reference_number', 'payment_date', 'partial_payment_amount']
        for field in allowed_fields:
            if field in data:
                if field == 'partial_payment_amount':
                    # Convert to float for partial payment amount
                    update_data[field] = float(data[field]) if data[field] else 0.0
                else:
                    update_data[field] = data[field]
        
        # If payment_status is being set to 'paid' or 'partial_paid', set payment_date if not provided
        if 'payment_status' in update_data and update_data['payment_status'] in ['paid', 'partial_paid']:
            if 'payment_date' not in update_data:
                update_data['payment_date'] = datetime.now()
        
        if len(update_data) == 1:  # Only updated_at
            return jsonify({'error': 'No valid fields to update'}), 400
        
        result = db.payouts.update_one({'_id': parse_object_id(id)}, {'$set': update_data})
        if result.modified_count:
            return jsonify({'message': 'Payout updated successfully'})
        return jsonify({'error': 'Payout not found'}), 404
    except Exception as e:
        logging.error(f"Error updating payout: {e}")
        return jsonify({'error': str(e)}), 500

# ==================== REPORTS API ====================

@app.route('/api/reports/summary', methods=['GET'])
def get_summary():
    try:
        summary = {
            'doctors': db.doctors.count_documents({}),
            'patients': db.patients.count_documents({}),
            'cases': db.cases.count_documents({}),
            'appointments': db.appointments.count_documents({}),
            'prescriptions': db.prescriptions.count_documents({})
        }
        return jsonify(summary)
    except Exception as e:
        logging.error(f"Error getting summary: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/payouts/export-excel', methods=['GET'])
def export_payouts_excel():
    try:
        # Get filter parameters
        case_id = request.args.get('case_id')
        doctor_id = request.args.get('doctor_id')
        payment_status = request.args.get('payment_status')
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        
        # Build query (same as get_payouts)
        query = {}
        if case_id:
            query['case_id'] = parse_object_id(case_id)
        if doctor_id:
            query['doctor_id'] = parse_object_id(doctor_id)
        if payment_status:
            query['payment_status'] = payment_status
        
        # Date range filtering
        if start_date and end_date:
            start_date_obj = datetime.strptime(start_date, '%Y-%m-%d')
            end_date_obj = datetime.strptime(end_date, '%Y-%m-%d')
            query['date_time'] = {
                '$gte': datetime.combine(start_date_obj.date(), datetime.min.time()),
                '$lte': datetime.combine(end_date_obj.date(), datetime.max.time())
            }
        
        # Get all payouts matching the query (no pagination for export)
        payouts = list(db.payouts.find(query).sort('date_time', -1))
        
        # Populate doctor names
        for payout in payouts:
            if 'doctor_id' in payout and payout['doctor_id']:
                doctor = db.doctors.find_one({'_id': payout['doctor_id']})
                if doctor:
                    payout['doctor_name'] = doctor.get('name', '')
        
        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Payouts Report"
        
        # Define header style
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Define column headers
        headers = [
            'Date & Time',
            'Case Number',
            'Patient Name',
            'Doctor Name',
            'OPD/IPD',
            'Total Charge Amount',
            'Doctor Charge Amount',
            'Payment Status',
            'Payment Date',
            'Payment Mode',
            'Payment Reference',
            'Partial Payment Amount',
            'Payment Comment'
        ]
        
        # Write headers
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # Write data rows
        for row_num, payout in enumerate(payouts, 2):
            date_time = payout.get('date_time', '')
            if date_time:
                if isinstance(date_time, datetime):
                    date_time = date_time.strftime('%Y-%m-%d %H:%M:%S')
                else:
                    date_time = str(date_time)
            
            payment_date = payout.get('payment_date', '')
            if payment_date:
                if isinstance(payment_date, datetime):
                    payment_date = payment_date.strftime('%Y-%m-%d')
                else:
                    payment_date = str(payment_date)
            
            # Set row color based on payment status
            row_fill = None
            if payout.get('payment_status') == 'paid':
                row_fill = PatternFill(start_color="d4edda", end_color="d4edda", fill_type="solid")
            elif payout.get('payment_status') == 'partial_paid':
                row_fill = PatternFill(start_color="cfe2ff", end_color="cfe2ff", fill_type="solid")
            elif payout.get('payment_status') in ['pending', 'cancelled']:
                row_fill = PatternFill(start_color="fff3cd", end_color="fff3cd", fill_type="solid")
            
            row_data = [
                date_time,
                payout.get('case_number', ''),
                payout.get('patient_name', ''),
                payout.get('doctor_name', ''),
                payout.get('case_type', ''),
                payout.get('total_charge_amount', 0),
                payout.get('doctor_charge_amount', 0),
                payout.get('payment_status', ''),
                payment_date,
                payout.get('payment_mode', ''),
                payout.get('payment_reference_number', ''),
                payout.get('partial_payment_amount', 0),
                payout.get('payment_comment', '')
            ]
            
            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                if row_fill:
                    cell.fill = row_fill
                # Format numeric columns
                if col_num in [6, 7, 12]:  # Amount columns
                    if isinstance(value, (int, float)):
                        cell.number_format = '#,##0.00'
        
        # Auto-adjust column widths
        for col_num, header in enumerate(headers, 1):
            column_letter = get_column_letter(col_num)
            max_length = len(str(header))
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=col_num, max_col=col_num):
                for cell in row:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
        
        # Add summary row
        summary_row = ws.max_row + 2
        ws.cell(row=summary_row, column=1, value="TOTAL").font = Font(bold=True)
        ws.cell(row=summary_row, column=6, value=sum(p.get('total_charge_amount', 0) for p in payouts)).font = Font(bold=True)
        ws.cell(row=summary_row, column=6).number_format = '#,##0.00'
        ws.cell(row=summary_row, column=7, value=sum(p.get('doctor_charge_amount', 0) for p in payouts)).font = Font(bold=True)
        ws.cell(row=summary_row, column=7).number_format = '#,##0.00'
        
        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Generate filename
        filename = f"payouts_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
    except Exception as e:
        logging.error(f"Error exporting payouts to Excel: {e}")
        return jsonify({'error': str(e)}), 500

# ==================== BILLING API ====================

@app.route('/api/billing/case/<case_id>', methods=['GET'])
def get_case_billing_details(case_id):
    """Get case details with charges for billing"""
    try:
        case = db.cases.find_one({'_id': parse_object_id(case_id)})
        if not case:
            return jsonify({'error': 'Case not found'}), 404
        
        # Populate patient
        if 'patient_id' in case:
            patient_id_obj = case['patient_id']
            if isinstance(patient_id_obj, str):
                patient_id_obj = parse_object_id(patient_id_obj)
            if patient_id_obj:
                patient = db.patients.find_one({'_id': patient_id_obj})
                if patient:
                    case['patient'] = serialize_doc(patient)
        
        # Get case charges (patient charges)
        case_charges = list(db.case_charges.find({'case_id': parse_object_id(case_id)}))
        for charge in case_charges:
            if 'charge_master_id' in charge:
                charge_master = db.charge_master.find_one({'_id': charge['charge_master_id']})
                if charge_master:
                    charge['charge_name'] = charge_master.get('name', '')
            if 'doctor_id' in charge:
                doctor = db.doctors.find_one({'_id': charge['doctor_id']})
                if doctor:
                    charge['doctor_name'] = doctor.get('name', '')
                    charge['doctor_specialization'] = doctor.get('specialization', '')
        
        # Get payments
        payments = list(db.payments.find({'case_id': parse_object_id(case_id)}).sort('payment_date', -1))
        
        # Calculate totals
        total_charges = sum(float(c.get('total_amount', 0) or 0) for c in case_charges)
        discount = float(case.get('discount', 0) or 0)
        total_after_discount = max(0, total_charges - discount)  # Ensure non-negative
        total_paid = sum(float(p.get('amount', 0) or 0) for p in payments)
        balance = total_after_discount - total_paid
        
        return jsonify({
            'case': serialize_doc(case),
            'charges': serialize_doc(case_charges),
            'payments': serialize_doc(payments),
            'total_charges': total_charges,
            'discount': discount,
            'total_after_discount': total_after_discount,
            'total_paid': total_paid,
            'balance': balance
        })
    except Exception as e:
        logging.error(f"Error getting case billing details: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/billing/case-charge/<charge_id>', methods=['PUT'])
def update_case_charge_for_billing(charge_id):
    """Update case charge amount for billing corrections"""
    try:
        data = request.get_json()
        
        # Get the existing charge to find case_id
        existing_charge = db.case_charges.find_one({'_id': parse_object_id(charge_id)})
        if not existing_charge:
            return jsonify({'error': 'Charge not found'}), 404
        
        case_id_obj = existing_charge.get('case_id')
        if case_id_obj:
            # Check if case is closed
            case = db.cases.find_one({'_id': case_id_obj})
            if case and case.get('status') == 'closed':
                return jsonify({'error': 'Cannot modify charges for a closed case'}), 400
        
        # Convert amounts to proper types
        if 'unit_amount' in data:
            data['unit_amount'] = float(data['unit_amount']) if data['unit_amount'] else 0.0
        if 'quantity' in data:
            data['quantity'] = int(data['quantity']) if data['quantity'] else 1
        if 'total_amount' in data:
            data['total_amount'] = float(data['total_amount']) if data['total_amount'] else 0.0
        elif 'unit_amount' in data and 'quantity' in data:
            # Calculate total_amount if not provided
            data['total_amount'] = float(data['unit_amount']) * int(data['quantity'])
        
        data['updated_at'] = datetime.now()
        result = db.case_charges.update_one({'_id': parse_object_id(charge_id)}, {'$set': data})
        if result.modified_count:
            return jsonify({'message': 'Charge updated successfully'})
        return jsonify({'error': 'Charge not found'}), 404
    except Exception as e:
        logging.error(f"Error updating case charge: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/billing/generate-pdf/<case_id>', methods=['GET'])
def generate_bill_pdf(case_id):
    """Generate PDF invoice for a case"""
    try:
        case = db.cases.find_one({'_id': parse_object_id(case_id)})
        if not case:
            return jsonify({'error': 'Case not found'}), 404
        
        # Get patient
        patient = None
        if 'patient_id' in case:
            patient_id_obj = case['patient_id']
            if isinstance(patient_id_obj, str):
                patient_id_obj = parse_object_id(patient_id_obj)
            if patient_id_obj:
                patient = db.patients.find_one({'_id': patient_id_obj})
        
        # Get charges
        case_charges = list(db.case_charges.find({'case_id': parse_object_id(case_id)}))
        for charge in case_charges:
            if 'charge_master_id' in charge:
                charge_master = db.charge_master.find_one({'_id': charge['charge_master_id']})
                if charge_master:
                    charge['charge_name'] = charge_master.get('name', '')
            if 'doctor_id' in charge:
                doctor = db.doctors.find_one({'_id': charge['doctor_id']})
                if doctor:
                    charge['doctor_name'] = doctor.get('name', '')
        
        # Get payments
        payments = list(db.payments.find({'case_id': parse_object_id(case_id)}).sort('payment_date', -1))
        
        # Calculate totals
        total_charges = sum(c.get('total_amount', 0) for c in case_charges)
        discount = case.get('discount', 0) or 0
        total_after_discount = max(0, total_charges - discount)  # Ensure non-negative
        total_paid = sum(p.get('amount', 0) for p in payments)
        balance = total_after_discount - total_paid
        
        # Create PDF
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=72, leftMargin=72, topMargin=72, bottomMargin=18)
        story = []
        
        # Styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#2563eb'),
            spaceAfter=30,
            alignment=1  # Center
        )
        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading2'],
            fontSize=14,
            textColor=colors.HexColor('#1f2937'),
            spaceAfter=12
        )
        
        # Title
        story.append(Paragraph("INVOICE / BILL", title_style))
        story.append(Spacer(1, 0.2*inch))
        
        # Hospital Info (you can customize this)
        hospital_info = [
            ["Life Plus Hospital"],
            ["Hospital Management System"],
            [f"Bill Date: {datetime.now().strftime('%d-%m-%Y %H:%M')}"]
        ]
        hospital_table = Table(hospital_info, colWidths=[4*inch])
        hospital_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (0, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (0, 0), 16),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ]))
        story.append(hospital_table)
        story.append(Spacer(1, 0.3*inch))
        
        # Patient and Case Info
        patient_info = []
        if patient:
            patient_info.append(["Patient Name:", patient.get('name', '')])
            patient_info.append(["Phone:", patient.get('phone', '')])
            patient_info.append(["Email:", patient.get('email', '')])
            if 'address' in patient:
                patient_info.append(["Address:", patient.get('address', '')])
        patient_info.append(["Case Number:", case.get('case_number', '')])
        patient_info.append(["Case Type:", case.get('case_type', '')])
        if case.get('admission_date'):
            patient_info.append(["Admission Date:", case.get('admission_date').strftime('%d-%m-%Y') if isinstance(case.get('admission_date'), datetime) else str(case.get('admission_date'))])
        
        patient_table = Table(patient_info, colWidths=[2*inch, 4*inch])
        patient_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#f3f4f6')),
            ('TEXTCOLOR', (0, 0), (0, -1), colors.HexColor('#374151')),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#e5e7eb')),
        ]))
        story.append(patient_table)
        story.append(Spacer(1, 0.3*inch))
        
        # Charges Table
        story.append(Paragraph("Charges Details", heading_style))
        charge_data = [["Date", "Charge Name", "Doctor", "Qty", "Unit Amount", "Total Amount"]]
        for charge in case_charges:
            charge_date = ''
            if charge.get('created_at'):
                charge_date = charge['created_at'].strftime('%d-%m-%Y') if isinstance(charge['created_at'], datetime) else str(charge['created_at'])[:10]
            charge_data.append([
                charge_date,
                charge.get('charge_name', ''),
                charge.get('doctor_name', ''),
                str(charge.get('quantity', 1)),
                f"{charge.get('unit_amount', 0):.2f}",
                f"{charge.get('total_amount', 0):.2f}"
            ])
        
        charge_table = Table(charge_data, colWidths=[1*inch, 2*inch, 1.5*inch, 0.5*inch, 1*inch, 1*inch])
        charge_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2563eb')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('ALIGN', (3, 0), (5, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 11),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#e5e7eb')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f9fafb')]),
        ]))
        story.append(charge_table)
        story.append(Spacer(1, 0.2*inch))
        
        # Payments Table
        if payments:
            story.append(Paragraph("Payment History", heading_style))
            payment_data = [["Date", "Amount", "Mode", "Reference", "Notes"]]
            for payment in payments:
                payment_date = ''
                if payment.get('payment_date'):
                    payment_date = payment['payment_date'].strftime('%d-%m-%Y') if isinstance(payment['payment_date'], datetime) else str(payment['payment_date'])[:10]
                payment_data.append([
                    payment_date,
                    f"{payment.get('amount', 0):.2f}",
                    payment.get('payment_mode', ''),
                    payment.get('payment_reference_number', ''),
                    payment.get('notes', '')
                ])
            
            payment_table = Table(payment_data, colWidths=[1.2*inch, 1*inch, 1*inch, 1.2*inch, 2.6*inch])
            payment_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#10b981')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 11),
                ('FONTSIZE', (0, 1), (-1, -1), 9),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
                ('TOPPADDING', (0, 0), (-1, -1), 8),
                ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#e5e7eb')),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f0fdf4')]),
            ]))
            story.append(payment_table)
            story.append(Spacer(1, 0.2*inch))
        
        # Summary
        story.append(Paragraph("Bill Summary", heading_style))
        summary_data = [
            ["Total Charges:", f"₹ {total_charges:.2f}"]
        ]
        if discount > 0:
            summary_data.append(["Discount:", f"-₹ {discount:.2f}"])
            summary_data.append(["Total After Discount:", f"₹ {total_after_discount:.2f}"])
        summary_data.extend([
            ["Total Paid:", f"₹ {total_paid:.2f}"],
            ["Balance Amount:", f"₹ {balance:.2f}"]
        ])
        summary_table = Table(summary_data, colWidths=[3*inch, 3*inch])
        # Calculate balance row index (last row)
        balance_row_idx = len(summary_data) - 1
        table_style = [
            ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#f3f4f6')),
            ('TEXTCOLOR', (0, 0), (0, -1), colors.HexColor('#374151')),
            ('ALIGN', (0, 0), (0, -1), 'LEFT'),
            ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTNAME', (1, 0), (1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 12),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
            ('TOPPADDING', (0, 0), (-1, -1), 10),
            ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#e5e7eb')),
            ('BACKGROUND', (0, balance_row_idx), (1, balance_row_idx), colors.HexColor('#fef3c7')),
        ]
        # Add discount row styling if discount exists
        if discount > 0:
            discount_row_idx = 1  # Discount is second row
            table_style.append(('TEXTCOLOR', (0, discount_row_idx), (1, discount_row_idx), colors.HexColor('#f59e0b')))
        summary_table.setStyle(TableStyle(table_style))
        story.append(summary_table)
        
        # Build PDF
        doc.build(story)
        buffer.seek(0)
        
        filename = f"Bill_{case.get('case_number', case_id)}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        
        return send_file(
            buffer,
            mimetype='application/pdf',
            as_attachment=True,
            download_name=filename
        )
    except Exception as e:
        logging.error(f"Error generating bill PDF: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/billing/discount/<case_id>', methods=['PUT'])
def update_case_discount(case_id):
    """Update discount for a case"""
    try:
        data = request.get_json()
        discount = float(data.get('discount', 0)) if data.get('discount') else 0.0
        
        # Ensure discount is non-negative
        if discount < 0:
            return jsonify({'error': 'Discount cannot be negative'}), 400
        
        # Check if case is closed
        case = db.cases.find_one({'_id': parse_object_id(case_id)})
        if not case:
            return jsonify({'error': 'Case not found'}), 404
        
        if case.get('status') == 'closed':
            return jsonify({'error': 'Cannot modify discount for a closed case'}), 400
        
        result = db.cases.update_one(
            {'_id': parse_object_id(case_id)},
            {'$set': {'discount': discount, 'updated_at': datetime.now()}}
        )
        if result.modified_count or result.matched_count:
            return jsonify({'message': 'Discount updated successfully', 'discount': discount})
        return jsonify({'error': 'Case not found'}), 404
    except Exception as e:
        logging.error(f"Error updating discount: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/billing/close-case/<case_id>', methods=['PUT'])
def close_case(case_id):
    """Close a case and mark it as billed - only if fully paid"""
    try:
        case = db.cases.find_one({'_id': parse_object_id(case_id)})
        if not case:
            return jsonify({'error': 'Case not found'}), 404
        
        # Check if case is already closed
        if case.get('status') == 'closed':
            return jsonify({'error': 'Case is already closed'}), 400
        
        # Calculate balance to ensure case is fully paid
        case_charges = list(db.case_charges.find({'case_id': parse_object_id(case_id)}))
        payments = list(db.payments.find({'case_id': parse_object_id(case_id)}))
        
        total_charges = sum(float(c.get('total_amount', 0) or 0) for c in case_charges)
        discount = float(case.get('discount', 0) or 0)
        total_after_discount = max(0, total_charges - discount)
        total_paid = sum(float(p.get('amount', 0) or 0) for p in payments)
        balance = total_after_discount - total_paid
        
        # Only allow closing if balance is zero (fully paid)
        if abs(balance) > 0.01:  # Allow small floating point differences
            return jsonify({
                'error': f'Cannot close case. Outstanding balance: ₹ {balance:.2f}. Please ensure all payments are completed before closing the case.'
            }), 400
        
        result = db.cases.update_one(
            {'_id': parse_object_id(case_id)},
            {'$set': {
                'status': 'closed', 
                'closed_at': datetime.now(), 
                'billed_at': datetime.now(),
                'updated_at': datetime.now()
            }}
        )
        if result.modified_count:
            return jsonify({'message': 'Case closed successfully'})
        return jsonify({'error': 'Case not found'}), 404
    except Exception as e:
        logging.error(f"Error closing case: {e}")
        return jsonify({'error': str(e)}), 500

# ==================== AUTHENTICATION & AUTHORIZATION ====================

def hash_password(password):
    """Hash password using SHA256"""
    return hashlib.sha256(password.encode()).hexdigest()

def check_permission(user, module, action):
    """Check if user has permission for module and action (view/edit/delete)"""
    if not user:
        return False
    
    # Admin user has full access
    if user.get('username') == 'sunilsahu':
        return True
    
    permissions = user.get('permissions', {})
    module_perms = permissions.get(module, {})
    
    # Check if user has the specific action permission
    if action == 'view':
        return module_perms.get('view', False)
    elif action == 'edit':
        return module_perms.get('edit', False)
    elif action == 'delete':
        return module_perms.get('delete', False)
    
    return False

def log_activity(user_id, username, action, module, details=None):
    """Log user activity for audit trail"""
    try:
        activity = {
            'user_id': parse_object_id(user_id) if user_id else None,
            'username': username,
            'action': action,  # 'view', 'create', 'update', 'delete', 'login', 'logout'
            'module': module,  # 'doctors', 'patients', 'cases', etc.
            'details': details or {},
            'ip_address': request.remote_addr,
            'timestamp': datetime.now()
        }
        db.activity_logs.insert_one(activity)
    except Exception as e:
        logging.error(f"Error logging activity: {e}")

# Initialize admin user if not exists
def initialize_admin_user():
    """Create admin user 'sunilsahu' if it doesn't exist"""
    try:
        admin = db.users.find_one({'username': 'sunilsahu'})
        if not admin:
            admin_user = {
                'username': 'sunilsahu',
                'password': hash_password('admin123'),  # Default password - should be changed
                'full_name': 'Admin User',
                'email': 'admin@hospital.com',
                'role': 'admin',
                'is_active': True,
                'permissions': {},  # Empty means full access for admin
                'created_at': datetime.now(),
                'created_by': None
            }
            db.users.insert_one(admin_user)
            logging.info("Admin user 'sunilsahu' created with default password 'admin123'")
    except Exception as e:
        logging.error(f"Error initializing admin user: {e}")

# Initialize admin on startup
initialize_admin_user()

# ==================== AUTHENTICATION API ====================

@app.route('/api/auth/login', methods=['POST'])
def login():
    """User login endpoint"""
    try:
        data = request.get_json()
        username = data.get('username', '').strip()
        password = data.get('password', '')
        
        if not username or not password:
            return jsonify({'error': 'Username and password are required'}), 400
        
        # Find user
        user = db.users.find_one({'username': username, 'is_active': True})
        if not user:
            log_activity(None, username, 'login_failed', 'auth', {'reason': 'user_not_found'})
            return jsonify({'error': 'Invalid username or password'}), 401
        
        # Check password
        hashed_password = hash_password(password)
        if user.get('password') != hashed_password:
            log_activity(str(user.get('_id')), username, 'login_failed', 'auth', {'reason': 'invalid_password'})
            return jsonify({'error': 'Invalid username or password'}), 401
        
        # Create session
        user_data = serialize_doc(user)
        user_data.pop('password', None)  # Remove password from response
        
        session['user_id'] = str(user.get('_id'))
        session['username'] = username
        session['role'] = user.get('role', 'user')
        
        # Log successful login
        log_activity(str(user.get('_id')), username, 'login', 'auth')
        
        return jsonify({
            'message': 'Login successful',
            'user': user_data
        })
    except Exception as e:
        logging.error(f"Error during login: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/auth/logout', methods=['POST'])
def logout():
    """User logout endpoint"""
    try:
        user_id = request.headers.get('X-User-Id')
        username = request.headers.get('X-Username', 'unknown')
        
        if user_id:
            log_activity(user_id, username, 'logout', 'auth')
        
        return jsonify({'message': 'Logout successful'})
    except Exception as e:
        logging.error(f"Error during logout: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/auth/check', methods=['GET'])
def check_auth():
    """Check if user is authenticated"""
    try:
        user_id = request.headers.get('X-User-Id')
        if not user_id:
            return jsonify({'authenticated': False}), 401
        
        user = db.users.find_one({'_id': parse_object_id(user_id), 'is_active': True})
        if not user:
            return jsonify({'authenticated': False}), 401
        
        user_data = serialize_doc(user)
        user_data.pop('password', None)
        return jsonify({'authenticated': True, 'user': user_data})
    except Exception as e:
        logging.error(f"Error checking auth: {e}")
        return jsonify({'authenticated': False}), 401

# ==================== USERS MANAGEMENT API ====================

@app.route('/api/users', methods=['GET'])
def get_users():
    """Get all users (admin only)"""
    try:
        user_id = request.headers.get('X-User-Id')
        if not user_id:
            return jsonify({'error': 'Unauthorized'}), 401
        
        user = db.users.find_one({'_id': parse_object_id(user_id)})
        if not user or user.get('username') != 'sunilsahu':
            return jsonify({'error': 'Access denied. Admin only.'}), 403
        
        users = list(db.users.find({}).sort('created_at', -1))
        for u in users:
            u.pop('password', None)  # Remove passwords
        
        log_activity(user_id, user.get('username'), 'view', 'users')
        return jsonify(serialize_doc(users))
    except Exception as e:
        logging.error(f"Error getting users: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/users', methods=['POST'])
def create_user():
    """Create new user (admin only)"""
    try:
        user_id = request.headers.get('X-User-Id')
        if not user_id:
            return jsonify({'error': 'Unauthorized'}), 401
        
        admin = db.users.find_one({'_id': parse_object_id(user_id)})
        if not admin or admin.get('username') != 'sunilsahu':
            return jsonify({'error': 'Access denied. Admin only.'}), 403
        
        data = request.get_json()
        username = data.get('username', '').strip()
        password = data.get('password', '')
        
        if not username or not password:
            return jsonify({'error': 'Username and password are required'}), 400
        
        # Check if username exists
        existing = db.users.find_one({'username': username})
        if existing:
            return jsonify({'error': 'Username already exists'}), 400
        
        # Create user
        new_user = {
            'username': username,
            'password': hash_password(password),
            'full_name': data.get('full_name', ''),
            'email': data.get('email', ''),
            'role': 'user',
            'is_active': data.get('is_active', True),
            'permissions': data.get('permissions', {}),
            'created_at': datetime.now(),
            'created_by': parse_object_id(user_id)
        }
        
        result = db.users.insert_one(new_user)
        new_user['_id'] = result.inserted_id
        new_user.pop('password', None)
        
        log_activity(user_id, admin.get('username'), 'create', 'users', {'target_user': username})
        return jsonify({'id': str(result.inserted_id), 'message': 'User created successfully', 'user': serialize_doc(new_user)}), 201
    except Exception as e:
        logging.error(f"Error creating user: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/users/<id>', methods=['PUT'])
def update_user(id):
    """Update user (admin only)"""
    try:
        user_id = request.headers.get('X-User-Id')
        if not user_id:
            return jsonify({'error': 'Unauthorized'}), 401
        
        admin = db.users.find_one({'_id': parse_object_id(user_id)})
        if not admin or admin.get('username') != 'sunilsahu':
            return jsonify({'error': 'Access denied. Admin only.'}), 403
        
        data = request.get_json()
        update_data = {}
        
        if 'full_name' in data:
            update_data['full_name'] = data['full_name']
        if 'email' in data:
            update_data['email'] = data['email']
        if 'is_active' in data:
            update_data['is_active'] = data['is_active']
        if 'permissions' in data:
            update_data['permissions'] = data['permissions']
        if 'password' in data and data['password']:
            update_data['password'] = hash_password(data['password'])
        
        update_data['updated_at'] = datetime.now()
        
        result = db.users.update_one({'_id': parse_object_id(id)}, {'$set': update_data})
        if result.modified_count:
            target_user = db.users.find_one({'_id': parse_object_id(id)})
            target_user.pop('password', None)
            log_activity(user_id, admin.get('username'), 'update', 'users', {'target_user_id': id})
            return jsonify({'message': 'User updated successfully', 'user': serialize_doc(target_user)})
        return jsonify({'error': 'User not found'}), 404
    except Exception as e:
        logging.error(f"Error updating user: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/users/<id>', methods=['DELETE'])
def delete_user(id):
    """Delete user (admin only)"""
    try:
        user_id = request.headers.get('X-User-Id')
        if not user_id:
            return jsonify({'error': 'Unauthorized'}), 401
        
        admin = db.users.find_one({'_id': parse_object_id(user_id)})
        if not admin or admin.get('username') != 'sunilsahu':
            return jsonify({'error': 'Access denied. Admin only.'}), 403
        
        # Prevent deleting admin user
        target_user = db.users.find_one({'_id': parse_object_id(id)})
        if target_user and target_user.get('username') == 'sunilsahu':
            return jsonify({'error': 'Cannot delete admin user'}), 400
        
        result = db.users.delete_one({'_id': parse_object_id(id)})
        if result.deleted_count:
            log_activity(user_id, admin.get('username'), 'delete', 'users', {'target_user_id': id})
            return jsonify({'message': 'User deleted successfully'})
        return jsonify({'error': 'User not found'}), 404
    except Exception as e:
        logging.error(f"Error deleting user: {e}")
        return jsonify({'error': str(e)}), 500

# ==================== ACTIVITY LOGS API ====================

@app.route('/api/activity-logs', methods=['GET'])
def get_activity_logs():
    """Get activity logs (admin only)"""
    try:
        user_id = request.headers.get('X-User-Id')
        if not user_id:
            return jsonify({'error': 'Unauthorized'}), 401
        
        admin = db.users.find_one({'_id': parse_object_id(user_id)})
        if not admin or admin.get('username') != 'sunilsahu':
            return jsonify({'error': 'Access denied. Admin only.'}), 403
        
        page = int(request.args.get('page', 1))
        limit = int(request.args.get('limit', 50))
        skip = (page - 1) * limit
        
        # Optional filters
        username_filter = request.args.get('username', '').strip()
        module_filter = request.args.get('module', '').strip()
        action_filter = request.args.get('action', '').strip()
        
        query = {}
        if username_filter:
            query['username'] = {'$regex': username_filter, '$options': 'i'}
        if module_filter:
            query['module'] = module_filter
        if action_filter:
            query['action'] = action_filter
        
        total = db.activity_logs.count_documents(query)
        logs = list(db.activity_logs.find(query).sort('timestamp', -1).skip(skip).limit(limit))
        
        log_activity(user_id, admin.get('username'), 'view', 'activity-logs')
        return jsonify({
            'logs': serialize_doc(logs),
            'total': total,
            'page': page,
            'limit': limit
        })
    except Exception as e:
        logging.error(f"Error getting activity logs: {e}")
        return jsonify({'error': str(e)}), 500

# ==================== DASHBOARD API ====================

@app.route('/api/dashboard/stats', methods=['GET'])
def get_dashboard_stats():
    """Get dashboard statistics"""
    try:
        # Total patients
        total_patients = db.patients.count_documents({})
        
        # Patients added this week
        week_ago = datetime.now() - timedelta(days=7)
        patients_this_week = db.patients.count_documents({'created_at': {'$gte': week_ago}})
        patients_trend = round((patients_this_week / max(total_patients - patients_this_week, 1)) * 100, 1) if total_patients > 0 else 0
        
        # Active cases (status is 'open' or 'Open' or doesn't exist)
        active_cases = db.cases.count_documents({
            '$or': [
                {'status': {'$regex': '^open$', '$options': 'i'}},
                {'status': {'$exists': False}}
            ]
        })
        
        # Adjust for IST (UTC+5:30)
        ist_offset = timedelta(hours=5, minutes=30)
        # today_start is server local time 00:00, but DB has UTC.
        today_start = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
        
        utc_today_start = today_start - ist_offset
        
        # Cases added today
        cases_today = db.cases.count_documents({'created_at': {'$gte': utc_today_start}})
        cases_trend = round((cases_today / max(active_cases - cases_today, 1)) * 100, 1) if active_cases > 0 else 0
        
        # Today's appointments
        today_appointments = db.appointments.count_documents({
            'appointment_date': {
                '$gte': utc_today_start,
                '$lt': utc_today_start + timedelta(days=1)
            }
        })
        
        # Yesterday's appointments for comparison
        utc_yesterday_start = utc_today_start - timedelta(days=1)
        yesterday_appointments = db.appointments.count_documents({
            'appointment_date': {
                '$gte': utc_yesterday_start,
                '$lt': utc_today_start
            }
        })
        appointments_trend = round(((today_appointments - yesterday_appointments) / max(yesterday_appointments, 1)) * 100, 1) if yesterday_appointments > 0 else 0
        
        # Revenue this month (Keep simplified for now or adjust purely for month)
        month_start = datetime.now().replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        payments_this_month = list(db.payments.find({'payment_date': {'$gte': month_start}}))
        revenue_this_month = sum(payment.get('amount', 0) for payment in payments_this_month)
        
        # Revenue last month for comparison
        if month_start.month == 1:
            last_month_start = month_start.replace(year=month_start.year - 1, month=12)
        else:
            last_month_start = month_start.replace(month=month_start.month - 1)
        
        payments_last_month = list(db.payments.find({
            'payment_date': {
                '$gte': last_month_start,
                '$lt': month_start
            }
        }))
        revenue_last_month = sum(payment.get('amount', 0) for payment in payments_last_month)
        revenue_trend = round(((revenue_this_month - revenue_last_month) / max(revenue_last_month, 1)) * 100, 1) if revenue_last_month > 0 else 0
        
        # Today's Collections
        payments_today = list(db.payments.find({
            'payment_date': {
                '$gte': utc_today_start,
                '$lt': utc_today_start + timedelta(days=1)
            }
        }))
        today_collections = sum(payment.get('amount', 0) for payment in payments_today)

        return jsonify({
            'total_patients': total_patients,
            'patients_trend': patients_trend,
            'active_cases': active_cases,
            'cases_trend': cases_trend,
            'today_appointments': today_appointments,
            'appointments_trend': appointments_trend,
            'revenue_this_month': revenue_this_month,
            'revenue_trend': revenue_trend,
            'today_collections': today_collections
        })
    except Exception as e:
        logging.error(f"Error getting dashboard stats: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/dashboard/financial-grid', methods=['GET'])
def get_financial_grid():
    """Get daily financial grid (collections vs payouts)"""
    try:
        # Default to last 10 days
        end_date_str = request.args.get('end_date')
        start_date_str = request.args.get('start_date')
        
        if end_date_str:
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d')
        else:
            end_date = datetime.now()
            
        if start_date_str:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d')
        else:
            start_date = end_date - timedelta(days=9) # Last 10 days including today
            
        # Ensure start/end cover full days
        start_dt = start_date.replace(hour=0, minute=0, second=0, microsecond=0)
        end_dt = end_date.replace(hour=23, minute=59, second=59, microsecond=999999)
        
        # Adjust for IST storage (IST is UTC+5:30)
        # We want to query based on IST days. 
        # IST Day Start (00:00 IST) is Prev Day 18:30 UTC.
        ist_offset = timedelta(hours=5, minutes=30)
        
        query_start = start_dt - ist_offset
        query_end = end_dt - ist_offset
        
        # 1. Collections (Payments)
        # Filter by payment_date
        payment_query = {
            'payment_date': {
                '$gte': query_start,
                '$lte': query_end
            }
        }
        payments = list(db.payments.find(payment_query))
        
        # Group by Date (YYYY-MM-DD)
        daily_collections = {}
        for p in payments:
            p_date = p.get('payment_date')
            if p_date:
                # Convert UTC to IST for display grouping
                local_date = p_date + ist_offset
                d_str = local_date.strftime('%Y-%m-%d')
                daily_collections[d_str] = daily_collections.get(d_str, 0) + (p.get('amount') or 0)
                
        # 2. Payouts
        # Filter by status 'paid' or 'partial_paid'
        payout_query = {
            'payment_status': {'$in': ['paid', 'partial_paid']},
            '$or': [
                {'payment_date': {'$gte': query_start, '$lte': query_end}},
                {'payment_date': None, 'date_time': {'$gte': query_start, '$lte': query_end}}
            ]
        }
        payouts = list(db.payouts.find(payout_query))
        
        daily_payouts = {}
        for p in payouts:
            # Determine effective data date
            eff_date = p.get('payment_date') or p.get('date_time')
            if not eff_date:
                continue
                
            # Filter again in python using query window
            if not (query_start <= eff_date <= query_end):
                continue
                
            # Convert to IST for display grouping
            local_date = eff_date + ist_offset
            d_str = local_date.strftime('%Y-%m-%d')
            
            amount = 0
            status = p.get('payment_status')
            
            if status == 'paid':
                try:
                     amount = float(p.get('doctor_charge_amount') or 0)
                except:
                     amount = 0
            elif status == 'partial_paid':
                try:
                    amount = float(p.get('partial_payment_amount') or 0)
                except:
                    amount = 0
            
            daily_payouts[d_str] = daily_payouts.get(d_str, 0) + amount
            
        # 3. Merge
        # Generate all dates in range
        grid_data = []
        curr = start_date
        while curr <= end_date:
            d_str = curr.strftime('%Y-%m-%d')
            coll = daily_collections.get(d_str, 0)
            pay = daily_payouts.get(d_str, 0)
            
            grid_data.append({
                'date': d_str,
                'collections': coll,
                'payouts': pay,
                'net': coll - pay
            })
            curr += timedelta(days=1)
            
        # Sort desc
        grid_data.sort(key=lambda x: x['date'], reverse=True)
        
        return jsonify(grid_data)

    except Exception as e:
        logging.error(f"Error getting financial grid: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/dashboard/financial-details', methods=['GET'])
def get_financial_details():
    """Get detailed financial transactions for a specific date"""
    try:
        date_str = request.args.get('date')
        if not date_str:
            return jsonify({'error': 'Date is required'}), 400
            
        target_date = datetime.strptime(date_str, '%Y-%m-%d')
        start_dt = target_date.replace(hour=0, minute=0, second=0, microsecond=0)
        end_dt = target_date.replace(hour=23, minute=59, second=59, microsecond=999999)
        
        # Adjust for IST (UTC+5:30)
        ist_offset = timedelta(hours=5, minutes=30)
        query_start = start_dt - ist_offset
        query_end = end_dt - ist_offset
        
        # 1. Collections (Payments)
        payment_query = {
            'payment_date': {
                '$gte': query_start,
                '$lte': query_end
            }
        }
        payments = list(db.payments.find(payment_query))
        
        collections_data = []
        for p in payments:
            patient_name = 'Unknown'
            if 'patient_id' in p:
                patient = db.patients.find_one({'_id': p['patient_id']})
                if patient:
                    patient_name = patient.get('name', 'Unknown')
                    
            case_number = 'N/A'
            if 'case_id' in p:
                case = db.cases.find_one({'_id': p['case_id']})
                if case:
                    case_number = case.get('case_number', 'N/A')
            
            collections_data.append({
                'patient_name': patient_name,
                'case_number': case_number,
                'amount': p.get('amount', 0),
                'mode': p.get('payment_mode', 'N/A')
            })
            
        # 2. Payouts
        payout_query = {
            'payment_status': {'$in': ['paid', 'partial_paid']},
            '$or': [
                {'payment_date': {'$gte': query_start, '$lte': query_end}},
                {'payment_date': None, 'date_time': {'$gte': query_start, '$lte': query_end}}
            ]
        }
        payouts = list(db.payouts.find(payout_query))
        
        payouts_data = []
        for p in payouts:
            # Re-verify effective date
            eff_date = p.get('payment_date') or p.get('date_time')
            if not eff_date or not (query_start <= eff_date <= query_end):
                continue
                
            doctor_name = 'Unknown'
            if 'doctor_id' in p:
                doctor = db.doctors.find_one({'_id': p['doctor_id']})
                if doctor:
                    doctor_name = doctor.get('name', 'Unknown')
            
            amount = 0
            status = p.get('payment_status')
            if status == 'paid':
                try: amount = float(p.get('doctor_charge_amount') or 0)
                except: amount = 0
            elif status == 'partial_paid':
                try: amount = float(p.get('partial_payment_amount') or 0)
                except: amount = 0
                
            payouts_data.append({
                'doctor_name': doctor_name,
                'type': status.replace('_', ' ').title(),
                'amount': amount
            })
            
        return jsonify({
            'date': date_str,
            'collections': collections_data,
            'payouts': payouts_data
        })

    except Exception as e:
        logging.error(f"Error getting financial details: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/dashboard/activity', methods=['GET'])
def get_dashboard_activity():
    """Get recent activity feed"""
    try:
        limit = int(request.args.get('limit', 10))
        activities = []
        
        # Recent patient registrations
        recent_patients = list(db.patients.find().sort('created_at', -1).limit(3))
        for patient in recent_patients:
            activities.append({
                'type': 'patient_registration',
                'icon': 'user',
                'message': f"Patient Registration (New) - {patient.get('name', 'Unknown')}",
                'timestamp': patient.get('created_at', datetime.now())
            })
        
        # Recent appointments
        recent_appointments = list(db.appointments.find().sort('created_at', -1).limit(3))
        for apt in recent_appointments:
            patient = db.patients.find_one({'_id': apt.get('patient_id')})
            patient_name = patient.get('name', 'Unknown') if patient else 'Unknown'
            activities.append({
                'type': 'appointment',
                'icon': 'calendar',
                'message': f"Appointment Scheduled - {patient_name}",
                'timestamp': apt.get('created_at', datetime.now())
            })
        
        # Recent payments
        recent_payments = list(db.payments.find().sort('payment_date', -1).limit(3))
        for payment in recent_payments:
            amount = payment.get('amount', 0)
            activities.append({
                'type': 'payment',
                'icon': 'money',
                'message': f"Payment Received - ₹{amount:,.2f}",
                'timestamp': payment.get('payment_date', datetime.now())
            })
        
        # Sort all activities by timestamp
        activities.sort(key=lambda x: x['timestamp'], reverse=True)
        
        # Limit to requested number
        activities = activities[:limit]
        
        # Serialize timestamps
        for activity in activities:
            if isinstance(activity['timestamp'], datetime):
                activity['timestamp'] = activity['timestamp'].isoformat()
        
        return jsonify({'activities': activities})
    except Exception as e:
        logging.error(f"Error getting dashboard activity: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/dashboard/upcoming-appointments', methods=['GET'])
def get_upcoming_appointments():
    """Get today's upcoming appointments"""
    try:
        # Get today's date range
        today_start = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
        today_end = today_start + timedelta(days=1)
        
        # Find today's appointments
        appointments = list(db.appointments.find({
            'appointment_date': {
                '$gte': today_start,
                '$lt': today_end
            }
        }).sort('appointment_time', 1))
        
        # Populate patient and doctor names
        for apt in appointments:
            if 'patient_id' in apt and apt['patient_id']:
                patient = db.patients.find_one({'_id': apt['patient_id']})
                if patient:
                    apt['patient_name'] = patient.get('name', 'Unknown')
                else:
                    apt['patient_name'] = 'Unknown'
            else:
                apt['patient_name'] = 'Unknown'
            
            if 'doctor_id' in apt and apt['doctor_id']:
                doctor = db.doctors.find_one({'_id': apt['doctor_id']})
                if doctor:
                    apt['doctor_name'] = doctor.get('name', 'Unknown')
                else:
                    apt['doctor_name'] = 'Unknown'
            else:
                apt['doctor_name'] = 'Unknown'
        
        return jsonify({
            'appointments': serialize_doc(appointments),
            'total': len(appointments)
        })
    except Exception as e:
        logging.error(f"Error getting upcoming appointments: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/case-studies', methods=['POST'])
def create_case_study():
    try:
        data = request.get_json()
        case_id = data.get('case_id')
        doctor_id = data.get('doctor_id')
        
        if is_case_closed(case_id):
             return jsonify({'error': 'Cannot add case studies to a closed case'}), 400
        title = data.get('study_title')
        details = data.get('details')
        
        if not all([case_id, title, details]):
            return jsonify({'error': 'Missing required fields'}), 400
            
        study_doc = {
            'case_id': parse_object_id(case_id),
            'doctor_id': parse_object_id(doctor_id) if doctor_id else None,
            'study_title': title,
            'details': details,
            'created_at': datetime.now(),
            'updated_at': datetime.now()
        }
        
        result = db.case_studies.insert_one(study_doc)
        return jsonify({'id': str(result.inserted_id), 'message': 'Case study added successfully'}), 201
    except Exception as e:
        logging.error(f"Error creating case study: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/case-studies/<id>', methods=['PUT'])
def update_case_study(id):
    try:
        data = request.get_json()
        title = data.get('study_title')
        details = data.get('details')
        
        existing = db.case_studies.find_one({'_id': parse_object_id(id)})
        if existing and is_case_closed(existing.get('case_id')):
            return jsonify({'error': 'Cannot update case study for a closed case'}), 400
        
        update_fields = {'updated_at': datetime.now()}
        if title:
            update_fields['study_title'] = title
        if details:
            update_fields['details'] = details
            
        result = db.case_studies.update_one(
            {'_id': parse_object_id(id)},
            {'$set': update_fields}
        )
        
        if result.matched_count == 0:
            return jsonify({'error': 'Case study not found'}), 404
            
        return jsonify({'message': 'Case study updated successfully'})
    except Exception as e:
        logging.error(f"Error updating case study: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/case-studies/<id>', methods=['DELETE'])
def delete_case_study(id):
    try:

        study_id = parse_object_id(id)
        existing = db.case_studies.find_one({'_id': study_id})
        if existing and is_case_closed(existing.get('case_id')):
             return jsonify({'error': 'Cannot delete case study for a closed case'}), 400

        result = db.case_studies.delete_one({'_id': study_id})
        if result.deleted_count == 0:
            return jsonify({'error': 'Case study not found'}), 404
        return jsonify({'message': 'Case study deleted successfully'})
    except Exception as e:
        logging.error(f"Error deleting case study: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/case-studies/<id>', methods=['GET'])
def get_case_study(id):
    try:
        study = db.case_studies.find_one({'_id': parse_object_id(id)})
        if not study:
            return jsonify({'error': 'Case study not found'}), 404
        return jsonify(serialize_doc(study))
    except Exception as e:
        logging.error(f"Error getting case study: {e}")
        return jsonify({'error': str(e)}), 500

if __name__ == '__main__':
    app.run(debug=True, port=5001)
# Trigger reload
